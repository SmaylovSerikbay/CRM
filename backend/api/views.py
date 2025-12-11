from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.exceptions import ValidationError
from django.conf import settings
from django.utils import timezone
from django.core.cache import cache
from django.http import HttpResponse, FileResponse
import sys
from django.contrib.auth.hashers import make_password, check_password
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
try:
    import qrcode
    from qrcode.image.pil import PilImage
    QRCODE_AVAILABLE = True
except ImportError:
    QRCODE_AVAILABLE = False
import requests
import random
import string
import io
import json
import os
import logging
from datetime import datetime
from django.db.models import Q

logger = logging.getLogger('api')
from .models import (
    User, ContingentEmployee, CalendarPlan, RouteSheet, DoctorExamination, Expertise,
    EmergencyNotification, HealthImprovementPlan, RecommendationTracking, Doctor,
    LaboratoryTest, FunctionalTest, Referral, PatientQueue, Contract, ContractHistory
)
from .serializers import (
    UserSerializer, ContingentEmployeeSerializer, CalendarPlanSerializer,
    RouteSheetSerializer, DoctorExaminationSerializer, ExpertiseSerializer,
    EmergencyNotificationSerializer, HealthImprovementPlanSerializer, RecommendationTrackingSerializer,
    DoctorSerializer, LaboratoryTestSerializer, FunctionalTestSerializer, ReferralSerializer,
    PatientQueueSerializer, ContractSerializer, ContractHistorySerializer
)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def _normalize_phone(self, phone):
        """Нормализует номер телефона для единообразного хранения в кэше"""
        if not phone:
            return phone
        # Убираем все нецифровые символы
        normalized = ''.join(filter(str.isdigit, str(phone)))
        # Если начинается с 8, заменяем на 7
        if normalized.startswith('8'):
            normalized = '7' + normalized[1:]
        # Если не начинается с 7 и есть цифры, добавляем 7
        if normalized and not normalized.startswith('7'):
            normalized = '7' + normalized
        return normalized

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def send_otp(self, request):
        logger.info(f"[OTP] Request received")
        logger.debug(f"[OTP] Request body: {request.body}")
        logger.debug(f"[OTP] Request data: {request.data}")
        logger.debug(f"[OTP] Content-Type: {request.content_type}")
        
        phone = request.data.get('phone')
        if not phone:
            logger.warning(f"[OTP] Error: Phone number is required")
            return Response({'error': 'Phone number is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Нормализуем номер телефона
        normalized_phone = self._normalize_phone(phone)
        logger.info(f"[OTP] Original phone: {phone}, Normalized: {normalized_phone}")

        # Generate OTP
        otp = ''.join(random.choices(string.digits, k=6))
        logger.info(f"[OTP] Generated OTP: {otp}")
        
        # Store OTP in cache (expires in 5 minutes) - используем нормализованный номер
        cache_key = f'otp_{normalized_phone}'
        cache.set(cache_key, otp, timeout=300)  # 5 minutes
        logger.info(f"[OTP] Stored OTP in cache with key: {cache_key}, timeout: 300 seconds")

        # Send OTP via Green API
        try:
            url = f"{settings.GREEN_API_URL}/waInstance{settings.GREEN_API_ID_INSTANCE}/sendMessage/{settings.GREEN_API_TOKEN}"
            # Используем нормализованный номер для WhatsApp (уже в формате 7XXXXXXXXXX)
            formatted_phone = normalized_phone
            
            payload = {
                "chatId": f"{formatted_phone}@c.us",
                "message": f"Ваш код подтверждения для входа в CRM: {otp}"
            }
            
            logger.info(f"[OTP] Sending to {formatted_phone}, URL: {url}")
            response = requests.post(url, json=payload, timeout=10)
            logger.info(f"[OTP] Response status: {response.status_code}, body: {response.text}")
            
            if response.status_code == 200:
                response_data = response.json()
                if response_data.get('idMessage'):
                    logger.info(f"[OTP] OTP sent successfully via Green API")
                    return Response({'message': 'OTP sent successfully'}, status=status.HTTP_200_OK)
                else:
                    logger.error(f"[OTP] Green API error: {response_data}")
                    return Response({'error': f'Green API error: {response_data.get("message", "Unknown error")}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                logger.error(f"[OTP] HTTP error: {response.status_code} - {response.text}")
                return Response({'error': f'Failed to send OTP: {response.text}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.exception(f"[OTP] Exception: {str(e)}")
            return Response({'error': f'Error sending OTP: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def verify_otp(self, request):
        try:
            logger.info(f"[OTP VERIFY] Request received")
            logger.info(f"[OTP VERIFY] Request body: {request.body}")
            logger.info(f"[OTP VERIFY] Request data: {request.data}")
            logger.info(f"[OTP VERIFY] Content-Type: {request.content_type}")
            
            phone = request.data.get('phone')
            otp = request.data.get('otp')
            
            logger.info(f"[OTP VERIFY] Phone: {phone}, OTP: {otp}")
            
            if not phone or not otp:
                error_msg = 'Phone and OTP are required'
                logger.warning(f"[OTP VERIFY] Error: {error_msg}")
                return Response({
                    'error': error_msg,
                    'details': {
                        'phone_provided': bool(phone),
                        'otp_provided': bool(otp)
                    }
                }, status=status.HTTP_400_BAD_REQUEST)

            # Нормализуем номер телефона (должен совпадать с форматом при отправке)
            normalized_phone = self._normalize_phone(phone)
            logger.info(f"[OTP VERIFY] Original phone: {phone}, Normalized: {normalized_phone}")

            # Get OTP from cache - используем нормализованный номер
            cache_key = f'otp_{normalized_phone}'
            stored_otp = cache.get(cache_key)
            logger.info(f"[OTP VERIFY] Cache key: {cache_key}, Stored OTP: {'***' if stored_otp else 'None'}")
            
            # Если OTP не найден, пробуем альтернативные форматы номера
            if not stored_otp:
                logger.warning(f"[OTP VERIFY] OTP not found in cache for key: {cache_key}")
                
                # Проверяем разные варианты формата номера
                test_phones = [
                    phone,
                    phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', ''),
                    '7' + phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '').lstrip('7'),
                    normalized_phone,
                ]
                
                found_key = None
                for test_phone in set(test_phones):
                    if not test_phone:
                        continue
                    test_normalized = self._normalize_phone(test_phone)
                    test_key = f'otp_{test_normalized}'
                    test_otp = cache.get(test_key)
                    if test_otp:
                        logger.info(f"[OTP VERIFY] Found OTP with alternative key: {test_key}")
                        found_key = test_key
                        stored_otp = test_otp
                        cache_key = test_key
                        break
                
                if not stored_otp:
                    logger.warning(f"[OTP VERIFY] OTP not found in cache after checking all formats")
                    return Response({
                        'error': 'OTP expired or invalid. Please request a new code.',
                        'details': {
                            'cache_key_checked': cache_key,
                            'phone_normalized': normalized_phone
                        }
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            if stored_otp != otp:
                logger.warning(f"[OTP VERIFY] OTP mismatch: stored={'***' if stored_otp else 'None'}, provided={otp}")
                return Response({
                    'error': 'Invalid OTP code. Please check and try again.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"[OTP VERIFY] OTP verified successfully")

            # Clear OTP from cache
            cache.delete(cache_key)

            # Get or create user - используем нормализованный номер
            try:
                user, created = User.objects.get_or_create(phone=normalized_phone, defaults={'username': normalized_phone})
                user.last_login_at = timezone.now()
                user.save()
                logger.info(f"[OTP VERIFY] User {'created' if created else 'found'}: {user.phone}")
            except Exception as e:
                logger.error(f"[OTP VERIFY] Error creating/finding user: {str(e)}")
                return Response({
                    'error': 'Error processing user account',
                    'details': str(e)
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            serializer = UserSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception(f"[OTP VERIFY] Unexpected error: {str(e)}")
            return Response({
                'error': 'An unexpected error occurred',
                'details': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def login_with_password(self, request):
        """Авторизация через телефон и пароль"""
        phone = request.data.get('phone')
        password = request.data.get('password')
        
        if not phone or not password:
            return Response({'error': 'Phone and password are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(phone=phone)
            
            # Проверяем, установлен ли пароль
            if not user.password:
                return Response({'error': 'Password not set for this user. Please use OTP login.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем пароль
            if check_password(password, user.password):
                user.last_login_at = timezone.now()
                user.save()
                
                serializer = UserSerializer(user)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response({'error': 'Invalid password'}, status=status.HTTP_401_UNAUTHORIZED)
                
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def set_password(self, request):
        """Установка/изменение пароля для пользователя"""
        phone = request.data.get('phone')
        new_password = request.data.get('new_password')
        old_password = request.data.get('old_password')  # Опционально, для изменения существующего пароля
        
        if not phone or not new_password:
            return Response({'error': 'Phone and new_password are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters long'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(phone=phone)
            
            # Если у пользователя уже есть пароль, требуем старый пароль для изменения
            if user.password and old_password:
                if not check_password(old_password, user.password):
                    return Response({'error': 'Invalid old password'}, status=status.HTTP_401_UNAUTHORIZED)
            
            # Устанавливаем новый пароль
            user.password = make_password(new_password)
            user.save()
            
            return Response({'message': 'Password set successfully'}, status=status.HTTP_200_OK)
                
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def complete_registration(self, request):
        phone = request.data.get('phone')
        role = request.data.get('role')
        registration_data = request.data.get('registration_data', {})
        clinic_role = request.data.get('clinic_role')
        password = request.data.get('password')  # Опциональный пароль при регистрации

        try:
            user = User.objects.get(phone=phone)
            user.role = role
            user.registration_data = registration_data
            user.registration_completed = True
            if role == 'clinic' and clinic_role:
                user.clinic_role = clinic_role
            
            # Если передан пароль при регистрации, устанавливаем его
            if password and len(password) >= 6:
                user.password = make_password(password)
            
            user.save()
            
            # ИСПРАВЛЕНИЕ: Если это работодатель, связываем существующие договоры с его БИНом
            if role == 'employer' and registration_data:
                bin_number = registration_data.get('bin') or registration_data.get('inn')
                if bin_number:
                    # Нормализуем БИН
                    bin_normalized = ''.join(str(bin_number).strip().split())
                    
                    # Находим все договоры с этим БИНом, которые еще не связаны с работодателем
                    unlinked_contracts = Contract.objects.filter(
                        employer_bin=bin_normalized,
                        employer__isnull=True
                    )
                    
                    # Связываем договоры с зарегистрированным работодателем
                    contracts_updated = unlinked_contracts.update(employer=user)
                    
                    logger.info(f"[REGISTRATION] Linked {contracts_updated} contracts to employer {user.id} with BIN {bin_normalized}")
                    
                    # Создаем записи в истории для каждого связанного договора
                    for contract in unlinked_contracts:
                        try:
                            ContractHistory.objects.create(
                                contract=contract,
                                action='employer_registered',
                                user=user,
                                user_role=user.role,
                                user_name=registration_data.get('name', ''),
                                comment=f'Работодатель зарегистрировался и договор автоматически связан с аккаунтом'
                            )
                        except Exception as e:
                            logger.error(f"Failed to create contract history: {str(e)}")
            
            serializer = UserSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def find_by_bin(self, request):
        """Поиск работодателя по БИН"""
        bin_number = request.query_params.get('bin')
        if not bin_number:
            return Response({'error': 'bin parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Нормализуем БИН (убираем пробелы, приводим к строке)
            bin_normalized = ''.join(str(bin_number).strip().split())
            
            # Ищем пользователя с указанным БИН в registration_data
            # Ищем по полям 'bin' и 'inn' (так как при регистрации может быть сохранено как 'inn')
            employers = User.objects.filter(role='employer')
            
            employer = None
            for emp in employers:
                reg_data = emp.registration_data or {}
                # Проверяем оба варианта: bin и inn
                emp_bin = str(reg_data.get('bin', '')).strip()
                emp_inn = str(reg_data.get('inn', '')).strip()
                
                # Нормализуем для сравнения
                emp_bin_normalized = ''.join(emp_bin.split())
                emp_inn_normalized = ''.join(emp_inn.split())
                
                if emp_bin_normalized == bin_normalized or emp_inn_normalized == bin_normalized:
                    employer = emp
                    break
            
            if employer:
                serializer = UserSerializer(employer)
                return Response({
                    'found': True,
                    'user': serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'found': False,
                    'message': 'Работодатель с таким БИН не найден. Будет создан новый договор.'
                }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def clinics(self, request):
        """Получить список всех клиник"""
        try:
            clinics = User.objects.filter(role='clinic', registration_completed=True)
            serializer = UserSerializer(clinics, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def find_clinic_by_bin(self, request):
        """Поиск клиники по БИН"""
        bin_number = request.query_params.get('bin')
        if not bin_number:
            return Response({'error': 'bin parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Нормализуем БИН (убираем пробелы, приводим к строке)
            bin_normalized = ''.join(str(bin_number).strip().split())
            
            # Ищем клинику с указанным БИН в registration_data
            clinics = User.objects.filter(role='clinic')
            
            clinic = None
            for cl in clinics:
                reg_data = cl.registration_data or {}
                # Проверяем оба варианта: bin и inn
                clinic_bin = str(reg_data.get('bin', '')).strip()
                clinic_inn = str(reg_data.get('inn', '')).strip()
                
                # Нормализуем для сравнения
                clinic_bin_normalized = ''.join(clinic_bin.split())
                clinic_inn_normalized = ''.join(clinic_inn.split())
                
                if clinic_bin_normalized == bin_normalized or clinic_inn_normalized == bin_normalized:
                    clinic = cl
                    break
            
            if clinic:
                serializer = UserSerializer(clinic)
                return Response({
                    'found': True,
                    'user': serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'found': False,
                    'message': 'Клиника с таким БИН не найдена. Можно отправить на субподряд незарегистрированной клинике.'
                }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_clinics_excel(self, request):
        """Экспорт списка клиник в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse
            import io
            
            # Получаем только клиники
            clinics = User.objects.filter(role='clinic').order_by('created_at')
            
            if not clinics.exists():
                return Response({'error': 'Нет данных для экспорта'}, status=status.HTTP_404_NOT_FOUND)
            
            # Создание Excel файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Клиники"
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовки
            headers = [
                'БИН/ИИН',
                'Название клиники',
                'Телефон',
                'Email',
                'Дата регистрации',
                'Статус регистрации'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            for row, clinic in enumerate(clinics, 2):
                # Получаем БИН из registration_data
                bin_iin = ''
                clinic_name = ''
                if clinic.registration_data:
                    bin_iin = clinic.registration_data.get('bin', '')
                    clinic_name = clinic.registration_data.get('clinic_name', '')
                
                data = [
                    bin_iin,
                    clinic_name or clinic.username,
                    clinic.phone,
                    clinic.email or '',
                    clinic.created_at.strftime('%d.%m.%Y %H:%M') if clinic.created_at else '',
                    'Завершена' if clinic.registration_completed else 'Не завершена'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 5:  # Дата
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 6:  # Статус
                        cell.alignment = Alignment(horizontal='center')
                        if value == 'Завершена':
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение в память
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Создание HTTP ответа
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="clinics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_employers_excel(self, request):
        """Экспорт списка работодателей в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse
            import io
            
            # Получаем только работодателей
            employers = User.objects.filter(role='employer').order_by('created_at')
            
            if not employers.exists():
                return Response({'error': 'Нет данных для экспорта'}, status=status.HTTP_404_NOT_FOUND)
            
            # Создание Excel файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Работодатели"
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовки
            headers = [
                'БИН/ИИН',
                'Название организации',
                'Телефон',
                'Email',
                'Дата регистрации',
                'Статус регистрации'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            for row, employer in enumerate(employers, 2):
                # Получаем БИН из registration_data
                bin_iin = ''
                company_name = ''
                if employer.registration_data:
                    bin_iin = employer.registration_data.get('bin', '')
                    company_name = employer.registration_data.get('company_name', '')
                
                data = [
                    bin_iin,
                    company_name or employer.username,
                    employer.phone,
                    employer.email or '',
                    employer.created_at.strftime('%d.%m.%Y %H:%M') if employer.created_at else '',
                    'Завершена' if employer.registration_completed else 'Не завершена'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 5:  # Дата
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 6:  # Статус
                        cell.alignment = Alignment(horizontal='center')
                        if value == 'Завершена':
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение в память
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Создание HTTP ответа
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="employers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ContingentEmployeeViewSet(viewsets.ModelViewSet):
    serializer_class = ContingentEmployeeSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        contract_id = self.request.query_params.get('contract_id')
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                
                # Если указан contract_id, фильтруем только по этому договору
                if contract_id:
                    try:
                        contract = Contract.objects.select_related('clinic', 'employer').get(id=contract_id)
                        # Проверяем права доступа к договору
                        if user.role == 'clinic' and contract.clinic == user:
                            return ContingentEmployee.objects.filter(contract=contract).select_related('user', 'contract')
                        elif user.role == 'employer' and (contract.employer == user or contract.employer_bin == user.registration_data.get('bin')):
                            return ContingentEmployee.objects.filter(contract=contract).select_related('user', 'contract')
                        else:
                            return ContingentEmployee.objects.none()
                    except Contract.DoesNotExist:
                        return ContingentEmployee.objects.none()
                
                # Если пользователь - клиника, возвращаем контингент всех работодателей И контингент, загруженный самой клиникой
                if user.role == 'clinic':
                    # Объединяем все варианты в один оптимизированный запрос
                    return ContingentEmployee.objects.filter(
                        Q(user__role='employer') | Q(user=user) | Q(contract__clinic=user)
                    ).select_related('user', 'contract').distinct()
                    print(f"DEBUG: Итоговый queryset: {queryset.count()} записей", file=sys.stderr)
                    
                    return queryset
                # Если пользователь - работодатель, возвращаем его контингент И контингент, загруженный клиникой по договору
                elif user.role == 'employer':
                    # Находим все договоры с этим работодателем (не только approved, но и активные)
                    user_bin = user.registration_data.get('bin') or user.registration_data.get('inn') if user.registration_data else None
                    allowed_statuses = ['approved', 'active', 'in_progress', 'sent', 'pending_approval']
                    contracts = Contract.objects.filter(
                        Q(employer=user) | (Q(employer_bin=user_bin) if user_bin else Q()),
                        status__in=allowed_statuses
                    ).select_related('clinic', 'employer')
                    
                    # Объединяем все варианты в один оптимизированный запрос
                    return ContingentEmployee.objects.filter(
                        Q(user=user) | Q(contract__in=contracts)
                    ).select_related('user', 'contract').distinct()
                else:
                    return ContingentEmployee.objects.filter(user=user).select_related('user', 'contract')
            except User.DoesNotExist:
                return ContingentEmployee.objects.none()
        return ContingentEmployee.objects.all().select_related('user', 'contract')

    @action(detail=False, methods=['get'])
    def by_contract_optimized(self, request):
        """Оптимизированное получение контингента по договору с пагинацией"""
        user_id = request.query_params.get('user_id')
        contract_id = request.query_params.get('contract_id')
        page_size = int(request.query_params.get('page_size', 50))  # По умолчанию 50 записей
        page = int(request.query_params.get('page', 1))
        
        if not user_id or not contract_id:
            return Response({'error': 'user_id и contract_id обязательны'}, status=400)
        
        try:
            user = User.objects.get(id=user_id)
            contract = Contract.objects.select_related('clinic', 'employer').get(id=contract_id)
            
            # Проверяем права доступа
            has_access = False
            if user.role == 'clinic' and contract.clinic == user:
                has_access = True
            elif user.role == 'employer' and (contract.employer == user or contract.employer_bin == user.registration_data.get('bin')):
                has_access = True
            
            if not has_access:
                return Response({'error': 'Нет доступа к этому договору'}, status=403)
            
            # Оптимизированный запрос с пагинацией
            import time
            start_time = time.time()
            
            # Упрощенный запрос без лишних JOIN - только нужные поля
            queryset = ContingentEmployee.objects.filter(
                contract=contract
            ).order_by('id')
            
            # Подсчитываем общее количество
            count_start = time.time()
            total_count = queryset.count()
            count_time = time.time() - count_start
            
            print(f"[PERFORMANCE] Contract {contract_id}: Count query took {count_time:.3f}s, found {total_count} records", file=sys.stderr)
            
            # Применяем пагинацию
            start = (page - 1) * page_size
            end = start + page_size
            
            pagination_start = time.time()
            paginated_queryset = queryset[start:end]
            pagination_time = time.time() - pagination_start
            
            # Сериализуем данные упрощенным способом без лишних запросов
            serialization_start = time.time()
            
            # Создаем упрощенные данные без дополнительных запросов
            simplified_data = []
            for emp in paginated_queryset:
                simplified_data.append({
                    'id': emp.id,
                    'name': emp.name,
                    'position': emp.position,
                    'department': emp.department,
                    'phone': emp.phone,
                    'birth_date': emp.birth_date,
                    'gender': emp.gender,
                    'harmful_factors': emp.harmful_factors,
                    'requires_examination': emp.requires_examination,
                    'last_examination_date': emp.last_examination_date,
                    'next_examination_date': emp.next_examination_date,
                    'total_experience_years': emp.total_experience_years,
                    'position_experience_years': emp.position_experience_years,
                    'notes': emp.notes,
                    'quarter': emp.quarter,
                    'contract': emp.contract_id,
                    'contract_number': contract.contract_number if contract else None,
                    'employer_name': contract.employer.registration_data.get('name') if contract and contract.employer and contract.employer.registration_data else None,
                })
            
            serialization_time = time.time() - serialization_start
            
            total_time = time.time() - start_time
            print(f"[PERFORMANCE] Contract {contract_id}: Pagination {pagination_time:.3f}s, Serialization {serialization_time:.3f}s, Total {total_time:.3f}s", file=sys.stderr)
            
            return Response({
                'results': simplified_data,
                'count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size,
                'has_next': end < total_count,
                'has_previous': page > 1
            })
            
        except (User.DoesNotExist, Contract.DoesNotExist):
            return Response({'error': 'Пользователь или договор не найден'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def counts_by_contract(self, request):
        """Быстрое получение счетчиков для договора без загрузки полных данных"""
        user_id = request.query_params.get('user_id')
        contract_id = request.query_params.get('contract_id')
        
        if not user_id or not contract_id:
            return Response({'error': 'user_id and contract_id are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            contract = Contract.objects.get(id=contract_id)
            
            # Проверяем права доступа к договору
            has_access = False
            if user.role == 'clinic' and contract.clinic == user:
                has_access = True
            elif user.role == 'employer' and (contract.employer == user or contract.employer_bin == user.registration_data.get('bin')):
                has_access = True
            
            if not has_access:
                return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
            
            # Быстрый подсчет без загрузки данных
            contingent_count = ContingentEmployee.objects.filter(contract=contract).count()
            plans_count = CalendarPlan.objects.filter(contract=contract).count()
            
            # Для маршрутных листов нужно найти ID сотрудников
            employee_ids = list(ContingentEmployee.objects.filter(contract=contract).values_list('id', flat=True))
            route_sheets_count = RouteSheet.objects.filter(patient_id__in=[str(eid) for eid in employee_ids]).count() if employee_ids else 0
            
            return Response({
                'contingent_count': contingent_count,
                'plans_count': plans_count,
                'route_sheets_count': route_sheets_count
            })
            
        except (User.DoesNotExist, Contract.DoesNotExist):
            return Response({'error': 'User or contract not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error getting counts for contract {contract_id}: {str(e)}")
            return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def generate_qr_code(self, request, pk=None):
        """Генерация QR-кода для сотрудника (работодатель)"""
        if not QRCODE_AVAILABLE:
            return Response({'error': 'QRCode library not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        try:
            employee = self.get_object()
            
            # Формируем данные для QR-кода сотрудника
            qr_data = {
                'type': 'employee',
                'employee_id': str(employee.id),
                'iin': employee.iin or '',
                'name': employee.name,
                'position': employee.position or '',
                'department': employee.department or '',
            }
            
            # Создаем QR-код
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(json.dumps(qr_data, ensure_ascii=False))
            qr.make(fit=True)
            
            # Создаем изображение
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Сохраняем в буфер
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='image/png')
            response['Content-Disposition'] = f'inline; filename="qr_employee_{employee.id}.png"'
            return response
            
        except ContingentEmployee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def find_by_qr(self, request):
        """Поиск сотрудника по QR-коду (клиника)"""
        qr_data_str = request.data.get('qr_data')
        
        if not qr_data_str:
            return Response({'error': 'QR data is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Декодируем данные QR-кода
            qr_data = json.loads(qr_data_str)
            
            if qr_data.get('type') != 'employee':
                return Response({'error': 'Invalid QR code type'}, status=status.HTTP_400_BAD_REQUEST)
            
            employee_id = qr_data.get('employee_id')
            iin = qr_data.get('iin')
            name = qr_data.get('name')
            
            # Ищем сотрудника
            employee = None
            if employee_id:
                try:
                    employee = ContingentEmployee.objects.get(id=employee_id)
                except ContingentEmployee.DoesNotExist:
                    pass
            
            if not employee and iin:
                employee = ContingentEmployee.objects.filter(iin=iin).first()
            
            if not employee and name:
                employee = ContingentEmployee.objects.filter(name__icontains=name).first()
            
            if not employee:
                return Response({
                    'error': 'Сотрудник не найден в базе данных',
                    'qr_data': qr_data
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ContingentEmployeeSerializer(employee)
            return Response({
                'employee': serializer.data,
                'found': True
            }, status=status.HTTP_200_OK)
            
        except json.JSONDecodeError:
            return Response({'error': 'Invalid QR code data format'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """Загрузка Excel файла со списком контингента"""
        try:
            import sys
            user_id = request.data.get('user_id')
            contract_id = request.data.get('contract_id')  # ID договора
            excel_file = request.FILES.get('file')
            replace_existing = request.data.get('replace_existing', 'false').lower() == 'true'  # Флаг замены существующих записей
            
            print(f"DEBUG UPLOAD: user_id={user_id}, contract_id={contract_id}, file={excel_file.name if excel_file else None}, replace_existing={replace_existing}", file=sys.stderr)
            
            if not excel_file:
                return Response({'error': 'Excel file is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if not user_id:
                return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'error': f'User with id {user_id} not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Проверяем наличие подтвержденного договора
            contract = None
            if contract_id:
                try:
                    contract = Contract.objects.get(id=contract_id)
                    # Проверяем статус договора в зависимости от роли пользователя
                    if user.role == 'employer':
                        # Для работодателей требуется подтвержденный договор
                        if contract.status != 'approved':
                            return Response({'error': 'Договор должен быть подтвержден перед загрузкой контингента'}, status=status.HTTP_400_BAD_REQUEST)
                    elif user.role == 'clinic':
                        # Для клиник разрешаем загрузку для активных договоров
                        allowed_statuses = ['approved', 'active', 'in_progress', 'sent', 'pending_approval']
                        if contract.status not in allowed_statuses:
                            return Response({'error': f'Загрузка контингента недоступна для договора со статусом "{contract.status}". Разрешенные статусы: {", ".join(allowed_statuses)}'}, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Проверяем, что пользователь связан с договором
                    if user.role == 'employer':
                        if contract.employer != user:
                            # Проверяем по БИН
                            reg_data = user.registration_data or {}
                            user_bin = reg_data.get('bin') or reg_data.get('inn')
                            if not user_bin or str(user_bin).strip() != str(contract.employer_bin or '').strip():
                                return Response({'error': 'Вы не являетесь стороной этого договора'}, status=status.HTTP_403_FORBIDDEN)
                    elif user.role == 'clinic':
                        if contract.clinic != user:
                            return Response({'error': 'Вы не являетесь стороной этого договора'}, status=status.HTTP_403_FORBIDDEN)
                except Contract.DoesNotExist:
                    return Response({'error': 'Договор не найден'}, status=status.HTTP_404_NOT_FOUND)
            else:
                # Если договор не указан, ищем активный договор между пользователем и его партнером
                if user.role == 'employer':
                    contracts = Contract.objects.filter(
                        Q(employer=user) | Q(employer_bin=user.registration_data.get('bin', '') if user.registration_data else ''),
                        status='approved'
                    ).first()
                    if contracts:
                        contract = contracts
                    else:
                        return Response({'error': 'Необходимо выбрать подтвержденный договор для загрузки контингента'}, status=status.HTTP_400_BAD_REQUEST)
                elif user.role == 'clinic':
                    # Для клиник ищем договоры с разрешенными статусами
                    allowed_statuses = ['approved', 'active', 'in_progress', 'sent', 'pending_approval']
                    contracts = Contract.objects.filter(clinic=user, status__in=allowed_statuses).first()
                    if contracts:
                        contract = contracts
                    else:
                        return Response({'error': f'Необходимо выбрать активный договор для загрузки контингента. Разрешенные статусы: {", ".join(allowed_statuses)}'}, status=status.HTTP_400_BAD_REQUEST)
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # ВАЛИДАЦИЯ СТРУКТУРЫ ФАЙЛА
            # Проверяем, что это наш шаблон по характерным признакам
            
            # 1. Проверяем заголовок документа в первой строке
            first_cell = worksheet['A1'].value
            if not first_cell or 'СПИСОК лиц, подлежащих обязательному медицинскому осмотру' not in str(first_cell):
                return Response({
                    'error': 'Загружаемый файл не соответствует шаблону',
                    'detail': 'Пожалуйста, скачайте последний действующий шаблон и заполните его'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # 2. Проверяем наличие ссылки на приказ во второй строке
            second_cell = worksheet['A2'].value
            if not second_cell or '№ ҚР ДСМ-131/2020' not in str(second_cell):
                return Response({
                    'error': 'Загружаемый файл не соответствует шаблону',
                    'detail': 'Пожалуйста, скачайте последний действующий шаблон и заполните его'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Пропускаем заголовки (первые 2-3 строки могут быть заголовками)
            # Ищем строку с заголовками колонок
            header_row = None
            for idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=5), start=1):
                row_values = [cell.value for cell in row if cell.value]
                if any('ФИО' in str(val) or '№ п/п' in str(val) for val in row_values):
                    header_row = idx
                    break
            
            if not header_row:
                return Response({
                    'error': 'Не найдены заголовки таблицы в шаблоне',
                    'detail': 'Пожалуйста, скачайте последний действующий шаблон и заполните его'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Маппинг колонок (поиск по заголовкам)
            column_map = {}
            header_cells = worksheet[header_row]
            for idx, cell in enumerate(header_cells, start=1):
                val = str(cell.value).lower() if cell.value else ''
                if '№ п/п' in val or 'номер' in val:
                    column_map['number'] = idx
                elif 'фио' in val or 'ф.и.о' in val:
                    column_map['name'] = idx
                elif 'дата рождения' in val:
                    column_map['birth_date'] = idx
                elif 'пол' in val:
                    column_map['gender'] = idx
                elif 'объект' in val or 'участок' in val:
                    column_map['department'] = idx
                elif 'должность' in val:
                    column_map['position'] = idx
                elif 'общий стаж' in val:
                    column_map['total_experience'] = idx
                elif 'стаж по должности' in val or 'стаж по занимаемой' in val:
                    column_map['position_experience'] = idx
                elif 'последний медосмотр' in val or 'дата последнего' in val:
                    column_map['last_examination'] = idx
                elif 'вредность' in val or 'профессиональная' in val:
                    column_map['harmful_factors'] = idx
                elif 'примечание' in val:
                    column_map['notes'] = idx
                elif 'иин' in val or 'иип' in val:
                    column_map['iin'] = idx
                elif 'телефон' in val:
                    column_map['phone'] = idx
                elif 'квартал' in val:
                    column_map['quarter'] = idx
            
            # 3. Проверяем наличие обязательных колонок
            required_columns = ['name', 'department', 'position']
            missing_columns = []
            for col in required_columns:
                if col not in column_map:
                    col_names = {
                        'name': 'ФИО',
                        'department': 'Объект/участок',
                        'position': 'Занимаемая должность'
                    }
                    missing_columns.append(col_names.get(col, col))
            
            if missing_columns:
                return Response({
                    'error': 'Загружаемый файл не соответствует шаблону',
                    'detail': f'Отсутствуют обязательные колонки: {", ".join(missing_columns)}. Пожалуйста, скачайте последний действующий шаблон и заполните его'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ВАЖНО: Проверяем наличие существующих записей
            existing_employees_count = ContingentEmployee.objects.filter(contract=contract).count()
            print(f"DEBUG UPLOAD START: Договор {contract.contract_number} (ID: {contract.id}) содержит {existing_employees_count} сотрудников ПЕРЕД загрузкой", file=sys.stderr)
            
            # Если есть существующие записи и не установлен флаг замены, предупреждаем пользователя
            if existing_employees_count > 0 and not replace_existing:
                return Response({
                    'warning': 'existing_data',
                    'message': f'В договоре уже есть {existing_employees_count} сотрудников. Загрузка нового файла заменит все существующие записи.',
                    'existing_count': existing_employees_count,
                    'action_required': 'confirm_replacement'
                }, status=status.HTTP_409_CONFLICT)
            
            # Если пользователь подтвердил замену, удаляем существующие записи
            if existing_employees_count > 0 and replace_existing:
                deleted_count = ContingentEmployee.objects.filter(contract=contract).count()
                ContingentEmployee.objects.filter(contract=contract).delete()
                print(f"DEBUG REPLACE: Удалено {deleted_count} существующих записей для замены", file=sys.stderr)
            
            created_employees = []
            skipped = 0
            skipped_reasons = {'duplicate': 0, 'no_name': 0}  # Убрали no_iin, так как ИИН теперь необязателен
            
            # Парсим данные начиная со строки после заголовков
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
                print(f"DEBUG ROW {row_idx}: Обрабатываем строку", file=sys.stderr)
                
                # Пропускаем пустые строки
                if not any(cell.value for cell in row):
                    print(f"DEBUG ROW {row_idx}: Пустая строка, пропускаем", file=sys.stderr)
                    continue
                
                # Извлекаем данные
                name = str(row[column_map.get('name', 2) - 1].value or '').strip()
                print(f"DEBUG ROW {row_idx}: ФИО = '{name}'", file=sys.stderr)
                
                if not name or name == 'None':
                    print(f"DEBUG ROW {row_idx}: Нет ФИО, пропускаем", file=sys.stderr)
                    skipped_reasons['no_name'] += 1
                    skipped += 1
                    continue
                
                # Парсим дату рождения
                birth_date = None
                birth_cell = row[column_map.get('birth_date', 3) - 1].value if column_map.get('birth_date') else None
                if birth_cell:
                    if isinstance(birth_cell, datetime):
                        birth_date = birth_cell.date()
                    elif isinstance(birth_cell, str):
                        try:
                            birth_date = datetime.strptime(birth_cell, '%d.%m.%Y').date()
                        except:
                            pass
                
                # Пол
                gender = None
                gender_cell = row[column_map.get('gender', 4) - 1].value if column_map.get('gender') else None
                if gender_cell:
                    gender_str = str(gender_cell).lower()
                    if 'муж' in gender_str or 'male' in gender_str:
                        gender = 'male'
                    elif 'жен' in gender_str or 'female' in gender_str:
                        gender = 'female'
                
                # Стаж
                total_exp = None
                pos_exp = None
                if column_map.get('total_experience'):
                    exp_val = row[column_map['total_experience'] - 1].value
                    if exp_val:
                        try:
                            total_exp = int(str(exp_val).replace('лет', '').replace('года', '').strip())
                        except:
                            pass
                
                if column_map.get('position_experience'):
                    exp_val = row[column_map['position_experience'] - 1].value
                    if exp_val:
                        try:
                            pos_exp = int(str(exp_val).replace('лет', '').replace('года', '').strip())
                        except:
                            pass
                
                # Дата последнего медосмотра
                last_exam = None
                exam_cell = row[column_map.get('last_examination', 9) - 1].value if column_map.get('last_examination') else None
                if exam_cell:
                    if isinstance(exam_cell, datetime):
                        last_exam = exam_cell.date()
                    elif isinstance(exam_cell, str):
                        try:
                            last_exam = datetime.strptime(exam_cell.replace('г', '').strip(), '%d.%m.%Y').date()
                        except:
                            pass
                
                # Вредные факторы
                harmful_factors = []
                factors_cell = row[column_map.get('harmful_factors', 10) - 1].value if column_map.get('harmful_factors') else None
                if factors_cell:
                    factors_str = str(factors_cell)
                    # Парсим список вредных факторов (могут быть через запятую или точку с запятой)
                    harmful_factors = [f.strip() for f in factors_str.replace(';', ',').split(',') if f.strip()]
                
                # ИИН (необязательное поле)
                iin_cell = row[column_map.get('iin', 5) - 1].value if column_map.get('iin') else None
                if not iin_cell:
                    # Пробуем найти ИИН в других колонках (может быть в другой позиции)
                    for col_idx in range(len(row)):
                        cell_val = str(row[col_idx].value or '').strip()
                        if len(cell_val) >= 10 and cell_val.isdigit():
                            iin_cell = cell_val
                            break
                
                iin = ''
                if iin_cell:
                    iin = str(iin_cell).strip()
                    # Очищаем ИИН от пробелов и других символов, оставляем только цифры
                    iin = ''.join(filter(str.isdigit, iin))
                    # Обрезаем до максимума 20 символов
                    iin = iin[:20] if len(iin) >= 10 else ''
                
                # Если ИИН не указан, генерируем уникальный идентификатор на основе ФИО и даты рождения
                if not iin:
                    import hashlib
                    unique_string = f"{name}_{birth_date or 'unknown'}_{row_idx}"
                    iin = hashlib.md5(unique_string.encode()).hexdigest()[:12]  # Используем первые 12 символов MD5
                
                # КРИТИЧЕСКИ ВАЖНО: Проверяем, существует ли уже сотрудник в рамках этого договора
                print(f"DEBUG ROW {row_idx}: Проверяем дубликаты для {name}, contract_id={contract.id if contract else 'None'}", file=sys.stderr)
                
                # Сначала получаем количество существующих записей в договоре
                existing_count = ContingentEmployee.objects.filter(contract=contract).count()
                print(f"DEBUG ROW {row_idx}: Существующих записей в договоре: {existing_count}", file=sys.stderr)
                
                existing = None
                
                # 1. Проверяем по ИИН в рамках этого договора (если ИИН валидный)
                if iin and len(iin) >= 10 and iin.isdigit():
                    existing = ContingentEmployee.objects.filter(contract=contract, iin=iin).first()
                    if existing:
                        print(f"DEBUG ROW {row_idx}: Найден дубликат по ИИН {iin} - {existing.name} (ID: {existing.id})", file=sys.stderr)
                
                # 2. Если не найден по ИИН, проверяем по ФИО + дате рождения в рамках этого договора
                if not existing:
                    existing = ContingentEmployee.objects.filter(
                        contract=contract, 
                        name__iexact=name,  # Используем iexact для регистронезависимого сравнения
                        birth_date=birth_date
                    ).first()
                    if existing:
                        print(f"DEBUG ROW {row_idx}: Найден дубликат по ФИО+дате рождения - {existing.name} (ID: {existing.id})", file=sys.stderr)
                
                # 3. Дополнительная проверка по ФИО (без даты рождения) для более строгого контроля
                if not existing:
                    name_duplicates = ContingentEmployee.objects.filter(
                        contract=contract, 
                        name__iexact=name
                    )
                    if name_duplicates.exists():
                        print(f"DEBUG ROW {row_idx}: Найдены записи с таким же ФИО: {[emp.id for emp in name_duplicates]}", file=sys.stderr)
                        # Если есть точное совпадение по ФИО, считаем дубликатом
                        existing = name_duplicates.first()
                
                if existing:
                    print(f"DEBUG ROW {row_idx}: ДУБЛИКАТ НАЙДЕН - {name} (существующий ID: {existing.id}), пропускаем", file=sys.stderr)
                    skipped_reasons['duplicate'] += 1
                    skipped += 1
                    continue
                
                print(f"DEBUG ROW {row_idx}: Дубликат не найден, создаем сотрудника", file=sys.stderr)
                
                # Нормализуем телефон для GreenAPI (формат 7XXXXXXXXXX)
                phone_raw = str(row[column_map.get('phone', 6) - 1].value or '').strip() if column_map.get('phone') else ''
                phone_normalized = ''
                if phone_raw:
                    # Убираем все символы кроме цифр
                    phone_digits = ''.join(filter(str.isdigit, phone_raw))
                    # Если начинается с 8, заменяем на 7
                    if phone_digits.startswith('8') and len(phone_digits) == 11:
                        phone_normalized = '7' + phone_digits[1:]
                    # Если начинается с 7 и длина 11 цифр
                    elif phone_digits.startswith('7') and len(phone_digits) == 11:
                        phone_normalized = phone_digits
                    # Если 10 цифр без кода страны, добавляем 7
                    elif len(phone_digits) == 10:
                        phone_normalized = '7' + phone_digits
                    # Иначе оставляем как есть (может быть международный номер)
                    else:
                        phone_normalized = phone_digits
                
                print(f"DEBUG CREATE: Создаем сотрудника {name} для пользователя {user.username} (ID: {user.id}) и договора {contract.contract_number if contract else 'None'} (ID: {contract.id if contract else 'None'})", file=sys.stderr)
                
                employee = ContingentEmployee.objects.create(
                    user=user,
                    contract=contract,
                    name=name,
                    birth_date=birth_date,
                    gender=gender,
                    department=str(row[column_map.get('department', 5) - 1].value or '').strip(),
                    position=str(row[column_map.get('position', 6) - 1].value or '').strip(),
                    total_experience_years=total_exp,
                    position_experience_years=pos_exp,
                    last_examination_date=last_exam,
                    harmful_factors=harmful_factors,
                    notes=str(row[column_map.get('notes', 11) - 1].value or '').strip() if column_map.get('notes') else '',
                    iin=iin,
                    phone=phone_normalized,
                    quarter=str(row[column_map.get('quarter', 1) - 1].value or '').strip() if column_map.get('quarter') else '',
                    requires_examination=True,
                )
                print(f"DEBUG CREATE: Сотрудник создан с ID: {employee.id}", file=sys.stderr)
                created_employees.append(employee)
            
            # ВАЖНО: Логируем состояние ПОСЛЕ загрузки
            final_employees_count = ContingentEmployee.objects.filter(contract=contract).count()
            print(f"DEBUG UPLOAD END: Договор {contract.contract_number} (ID: {contract.id}) содержит {final_employees_count} сотрудников ПОСЛЕ загрузки", file=sys.stderr)
            print(f"DEBUG UPLOAD SUMMARY: Было {existing_employees_count}, создано {len(created_employees)}, пропущено {skipped}, итого должно быть {existing_employees_count + len(created_employees)}, фактически {final_employees_count}", file=sys.stderr)
            
            serializer = ContingentEmployeeSerializer(created_employees, many=True)
            return Response({
                'created': len(created_employees),
                'skipped': skipped,
                'skipped_reasons': skipped_reasons,
                'employees': serializer.data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            import traceback
            import sys
            error_trace = traceback.format_exc()
            error_msg = str(e)
            print(f"Error in upload_excel: {error_msg}", file=sys.stderr)
            print(f"Traceback: {error_trace}", file=sys.stderr)
            return Response({
                'error': error_msg,
                'traceback': error_trace if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def harmful_factors(self, request):
        """Получение списка стандартных вредных факторов"""
        harmful_factors = [
            "п.1 «Работы, связанные с воздействием химических факторов»",
            "п.2 «Работы с канцерогенными веществами»",
            "п.3 «Работы с пестицидами и агрохимикатами»",
            "п.4 «Работы, связанные с воздействием биологических факторов»",
            "п.5 «Работы, выполняемые в условиях повышенного шума»",
            "п.6 «Работы, выполняемые в условиях вибрации»",
            "п.7 «Работы, выполняемые в условиях ионизирующего излучения»",
            "п.8 «Работы, выполняемые в условиях неионизирующих излучений»",
            "п.9 «Работы, выполняемые при повышенной или пониженной температуре воздуха»",
            "п.10 «Работы в замкнутых пространствах»",
            "п.11 «Работы на высоте»",
            "п.12 «Работы, связанные с подъемом и перемещением тяжестей»",
            "п.13 «Работы в ночное время»",
            "п.14 «Работа на ПК»",
            "п.15 «Работы, связанные с эмоциональным и умственным перенапряжением»",
            "п.16 «Работы, связанные с повышенной ответственностью»",
            "п.17 «Работы вахтовым методом»",
            "п.18 «Подземные работы»",
            "п.19 «Работы на транспорте»",
            "п.20 «Работы, связанные с воздействием пыли»",
            "п.21 «Работы с горюче-смазочными материалами»",
            "п.22 «Работы, связанные с воздействием нефти и нефтепродуктов»",
            "п.23 «Работы в условиях повышенной загазованности»",
            "п.24 «Работы в условиях недостатка кислорода»",
            "п.25 «Работы в условиях повышенной влажности»",
            "п.26 «Работы, связанные с виброинструментом»",
            "п.27 «Работы на конвейерах»",
            "п.28 «Работы на строительных площадках»",
            "п.29 «Работы в металлургическом производстве»",
            "п.30 «Работы в горнодобывающей промышленности»",
            "п.31 «Работы в деревообрабатывающем производстве»",
            "п.32 «Работы в текстильной и швейной промышленности»",
            "п.33 «Профессии и работы»"
        ]
        
        return Response({
            'harmful_factors': harmful_factors
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Скачивание шаблона Excel для списка контингента"""
        # Сначала пытаемся отдать заранее подготовленный статичный шаблон,
        # в котором уже настроены выпадающие списки, стили и т.п.
        # Это самый надежный способ, т.к. файл формируется самим пользователем в Excel.
        static_template_path = os.path.join(
            settings.BASE_DIR,
            "backend",
            "static",
            "templates",
            "contingent_template.xlsx",
        )
        if os.path.exists(static_template_path):
            # Комментарий (RU): если файл существует, отдаем его как есть
            return FileResponse(
                open(static_template_path, "rb"),
                as_attachment=True,
                filename="шаблон_список_контингента.xlsx",
            )

        # Если статичный файл не найден, генерируем шаблон программно (без сложных списков)
        # Создаем новый Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Список контингента"

        # Заголовок документа с текущим годом
        from datetime import datetime
        current_year = datetime.now().year
        ws.merge_cells('A1:K1')
        ws['A1'] = f'СПИСОК лиц, подлежащих обязательному медицинскому осмотру в {current_year} году'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:K2')
        ws['A2'] = 'согласно приказу и.о. Министра здравоохранения Республики Казахстан от 15 октября 2020 года № ҚР ДСМ-131/2020'
        ws['A2'].font = Font(size=10)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        # Подзаголовок квартала с выпадающим списком
        ws.merge_cells('H3:I3')
        ws['H3'] = '1 квартал'
        ws['H3'].font = Font(bold=True)
        ws['H3'].alignment = Alignment(horizontal='center')
        
        # Добавляем комментарий для ячейки квартала
        from openpyxl.comments import Comment
        quarter_comment = Comment(
            "📅 ВЫБОР КВАРТАЛА\n\n"
            "Кликните на эту ячейку и выберите квартал из выпадающего списка.\n\n"
            "Доступные варианты:\n"
            "• 1 квартал (январь-март)\n"
            "• 2 квартал (апрель-июнь)\n"
            "• 3 квартал (июль-сентябрь)\n"
            "• 4 квартал (октябрь-декабрь)\n\n"
            "⚠️ Выбранный квартал будет применен ко всему списку контингента",
            "Система"
        )
        quarter_comment.width = 350
        quarter_comment.height = 180
        ws['H3'].comment = quarter_comment
        
        # Добавляем выпадающий список для квартала
        quarter_validation = DataValidation(
            type="list",
            formula1='"1 квартал,2 квартал,3 квартал,4 квартал"',
            allow_blank=False,
            showDropDown=False,
            showInputMessage=True,
            showErrorMessage=True
        )
        quarter_validation.error = "❌ Выберите квартал из списка"
        quarter_validation.errorTitle = "Неверное значение"
        quarter_validation.prompt = "📋 Выберите квартал:\n• 1 квартал\n• 2 квартал\n• 3 квартал\n• 4 квартал"
        quarter_validation.promptTitle = "Выбор квартала"
        ws.add_data_validation(quarter_validation)
        quarter_validation.add('H3:I3')

        # Заголовки колонок
        headers = [
            '№ п/п',
            'ФИО',
            'Дата рождения',
            'Пол',
            'Телефон',
            'Объект или участок',
            'Занимаемая должность',
            'Общий стаж',
            'Стаж по занимаемой должности',
            'Дата последнего медосмотра',
            'Профессиональная вредность',
            'Примечание'
        ]

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Добавляем комментарии к заголовкам колонок с инструкциями
            if col_idx == 1:  # Колонка "№ п/п"
                from openpyxl.comments import Comment
                comment = Comment(
                    "🔢 ПОРЯДКОВЫЙ НОМЕР\n\n"
                    "Укажите порядковый номер сотрудника в списке.\n\n"
                    "Примеры:\n"
                    "• 1\n"
                    "• 2\n"
                    "• 3\n"
                    "• ...\n\n"
                    "ℹ️ Нумерация должна быть последовательной",
                    "Система"
                )
                comment.width = 300
                comment.height = 150
                cell.comment = comment
            elif col_idx == 2:  # Колонка "ФИО"
                from openpyxl.comments import Comment
                comment = Comment(
                    "👤 ФИО СОТРУДНИКА\n\n"
                    "Введите полное ФИО сотрудника.\n\n"
                    "Формат:\n"
                    "Фамилия Имя Отчество\n\n"
                    "Примеры:\n"
                    "• Иванов Иван Иванович\n"
                    "• Петрова Мария Петровна\n"
                    "• Сидоров Алексей Владимирович\n\n"
                    "⚠️ Это обязательное поле!",
                    "Система"
                )
                comment.width = 300
                comment.height = 160
                cell.comment = comment
            elif col_idx == 3:  # Колонка "Дата рождения"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📅 ФОРМАТ ДАТЫ\n\n"
                    "Введите дату в формате:\n"
                    "ДД.ММ.ГГГГ\n\n"
                    "Примеры:\n"
                    "• 29.03.1976\n"
                    "• 15.05.1985\n"
                    "• 01.01.1990\n\n"
                    "⚠️ Обязательно используйте точки между числами!",
                    "Система"
                )
                comment.width = 300
                comment.height = 150
                cell.comment = comment
            elif col_idx == 4:  # Колонка "Пол"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📋 ВЫПАДАЮЩИЙ СПИСОК\n\n"
                    "Кликните на ячейку ниже и выберите значение из списка:\n"
                    "• мужской\n"
                    "• женский\n\n"
                    "⚠️ Не вводите текст вручную!",
                    "Система"
                )
                comment.width = 300
                comment.height = 120
                cell.comment = comment
            elif col_idx == 5:  # Колонка "Телефон"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📱 ТЕЛЕФОН (необязательно)\n\n"
                    "Введите номер телефона сотрудника для отправки уведомлений в WhatsApp.\n\n"
                    "Формат (любой из вариантов):\n"
                    "• +77001234567\n"
                    "• 87001234567\n"
                    "• 77001234567\n"
                    "• 7001234567\n\n"
                    "⚠️ Номер будет автоматически преобразован в формат 7XXXXXXXXXX\n\n"
                    "ℹ️ Это поле можно оставить пустым",
                    "Система"
                )
                comment.width = 320
                comment.height = 170
                cell.comment = comment
            elif col_idx == 6:  # Колонка "Объект или участок"
                from openpyxl.comments import Comment
                comment = Comment(
                    "🏢 ОБЪЕКТ ИЛИ УЧАСТОК\n\n"
                    "Укажите место работы сотрудника:\n"
                    "• Название объекта\n"
                    "• Участок\n"
                    "• Отдел\n"
                    "• Цех\n\n"
                    "Примеры:\n"
                    "• ТОО \"Компания\" - Отдел продаж\n"
                    "• Производственный участок №1\n"
                    "• Административный корпус\n"
                    "• Цех металлообработки\n\n"
                    "⚠️ Это обязательное поле!",
                    "Система"
                )
                comment.width = 320
                comment.height = 180
                cell.comment = comment
            elif col_idx == 7:  # Колонка "Занимаемая должность"
                from openpyxl.comments import Comment
                comment = Comment(
                    "💼 ЗАНИМАЕМАЯ ДОЛЖНОСТЬ\n\n"
                    "Укажите должность сотрудника согласно штатному расписанию.\n\n"
                    "Примеры:\n"
                    "• Оператор станков с ЧПУ\n"
                    "• Главный бухгалтер\n"
                    "• Инженер-технолог\n"
                    "• Водитель погрузчика\n"
                    "• Менеджер по продажам\n"
                    "• Электромонтер\n\n"
                    "⚠️ Это обязательное поле!",
                    "Система"
                )
                comment.width = 320
                comment.height = 180
                cell.comment = comment
            elif col_idx == 8:  # Колонка "Общий стаж"
                from openpyxl.comments import Comment
                comment = Comment(
                    "🔢 ТОЛЬКО ЦИФРЫ\n\n"
                    "Введите количество лет общего стажа работы.\n\n"
                    "Примеры:\n"
                    "• 20\n"
                    "• 15\n"
                    "• 5\n\n"
                    "⚠️ Вводите только цифры без текста!",
                    "Система"
                )
                comment.width = 300
                comment.height = 140
                cell.comment = comment
            elif col_idx == 9:  # Колонка "Стаж по занимаемой должности"
                from openpyxl.comments import Comment
                comment = Comment(
                    "🔢 ТОЛЬКО ЦИФРЫ\n\n"
                    "Введите количество лет стажа по текущей должности.\n\n"
                    "Примеры:\n"
                    "• 18\n"
                    "• 10\n"
                    "• 3\n\n"
                    "⚠️ Вводите только цифры без текста!",
                    "Система"
                )
                comment.width = 300
                comment.height = 140
                cell.comment = comment
            elif col_idx == 10:  # Колонка "Дата последнего медосмотра"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📅 ФОРМАТ ДАТЫ\n\n"
                    "Введите дату в формате:\n"
                    "ДД.ММ.ГГГГ\n\n"
                    "Примеры:\n"
                    "• 22.01.2024\n"
                    "• 15.03.2023\n"
                    "• 01.12.2024\n\n"
                    "⚠️ Обязательно используйте точки между числами!",
                    "Система"
                )
                comment.width = 300
                comment.height = 150
                cell.comment = comment
            elif col_idx == 11:  # Колонка "Профессиональная вредность"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📋 ВЫПАДАЮЩИЙ СПИСОК\n\n"
                    "Кликните на ячейку ниже и выберите категорию из списка.\n\n"
                    "Доступно 33 категории согласно приказу № ҚР ДСМ-131/2020:\n"
                    "п.1 - Химические факторы\n"
                    "п.2 - Канцерогенные вещества\n"
                    "п.14 - Работа на ПК\n"
                    "п.33 - Профессии и работы\n"
                    "и другие...\n\n"
                    "⚠️ Не вводите текст вручную!",
                    "Система"
                )
                comment.width = 350
                comment.height = 180
                cell.comment = comment
            elif col_idx == 12:  # Колонка "Примечание"
                from openpyxl.comments import Comment
                comment = Comment(
                    "📝 ПРИМЕЧАНИЕ (необязательно)\n\n"
                    "Здесь можно указать дополнительную информацию о сотруднике.\n\n"
                    "Примеры:\n"
                    "• Работает по совместительству\n"
                    "• Временно отсутствует\n"
                    "• Находится в декретном отпуске\n"
                    "• Имеет медицинские противопоказания\n"
                    "• Любая другая важная информация\n\n"
                    "ℹ️ Это поле можно оставить пустым",
                    "Система"
                )
                comment.width = 320
                comment.height = 180
                cell.comment = comment

        # Пример данных
        example_data = [
            ['1', 'Иванов Иван Иванович', '29.03.1976', 'мужской', '77001234567', 'ТОО "Компания" - Отдел', 'Оператор', '20', '18', '22.01.2024', 'п.33 «Профессии и работы»', ''],
            ['2', 'Петрова Мария Петровна', '15.05.1985', 'женский', '', 'ТОО "Компания" - Офис', 'Бухгалтер', '15', '10', '24.01.2024', 'п.14 «Работа на ПК»', '']
        ]

        for row_idx, row_data in enumerate(example_data, start=header_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                # Устанавливаем текстовый формат для колонки телефона (E)
                if col_idx == 5:  # Колонка "Телефон"
                    cell.number_format = '@'  # Текстовый формат

        # Настройка ширины колонок
        column_widths = [8, 30, 15, 10, 15, 30, 25, 12, 25, 20, 40, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Высота строк
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[header_row].height = 40

        # --- Выпадающие списки ---
        # Используем именованные диапазоны на скрытых листах, чтобы Excel надёжно видел источники списков
        data_start_row = header_row + 1
        data_end_row = 500  # разумный лимит строк в шаблоне

        # 1) Пол: используем прямой список (более надежный способ)
        gender_validation = DataValidation(
            type="list",
            formula1='"мужской,женский"',  # Прямой список через запятую
            allow_blank=True,
            showDropDown=False,  # False показывает стрелку! (особенность openpyxl)
            showInputMessage=True,
            showErrorMessage=True
        )
        gender_validation.error = "❌ Пожалуйста, выберите пол из выпадающего списка"
        gender_validation.errorTitle = "Неверное значение"
        gender_validation.prompt = "📋 Кликните на стрелку ▼ справа и выберите:\n• мужской\n• женский"
        gender_validation.promptTitle = "Выбор пола"
        ws.add_data_validation(gender_validation)
        gender_validation.add(f"D{data_start_row}:D{data_end_row}")  # Колонка D - Пол
        
        # Создаем справочный лист для документации (необязательно для работы списка)
        gender_sheet = wb.create_sheet(title="Ref_Gender")
        gender_values = ["мужской", "женский"]
        for idx, gender in enumerate(gender_values, start=1):
            gender_sheet.cell(row=idx, column=1, value=gender)

        # 2) Профессиональная вредность: справочник на отдельном листе
        harmful_sheet = wb.create_sheet(title="Ref_Harm")

        # Комментарий (RU): наполняем справочник типовыми формулировками согласно приказу № ҚР ДСМ-131/2020
        harmful_factors = [
            "п.1 «Работы, связанные с воздействием химических факторов»",
            "п.2 «Работы с канцерогенными веществами»",
            "п.3 «Работы с пестицидами и агрохимикатами»",
            "п.4 «Работы, связанные с воздействием биологических факторов»",
            "п.5 «Работы, выполняемые в условиях повышенного шума»",
            "п.6 «Работы, выполняемые в условиях вибрации»",
            "п.7 «Работы, выполняемые в условиях ионизирующего излучения»",
            "п.8 «Работы, выполняемые в условиях неионизирующих излучений»",
            "п.9 «Работы, выполняемые при повышенной или пониженной температуре воздуха»",
            "п.10 «Работы в замкнутых пространствах»",
            "п.11 «Работы на высоте»",
            "п.12 «Работы, связанные с подъемом и перемещением тяжестей»",
            "п.13 «Работы в ночное время»",
            "п.14 «Работа на ПК»",
            "п.15 «Работы, связанные с эмоциональным и умственным перенапряжением»",
            "п.16 «Работы, связанные с повышенной ответственностью»",
            "п.17 «Работы вахтовым методом»",
            "п.18 «Подземные работы»",
            "п.19 «Работы на транспорте»",
            "п.20 «Работы, связанные с воздействием пыли»",
            "п.21 «Работы с горюче-смазочными материалами»",
            "п.22 «Работы, связанные с воздействием нефти и нефтепродуктов»",
            "п.23 «Работы в условиях повышенной загазованности»",
            "п.24 «Работы в условиях недостатка кислорода»",
            "п.25 «Работы в условиях повышенной влажности»",
            "п.26 «Работы, связанные с виброинструментом»",
            "п.27 «Работы на конвейерах»",
            "п.28 «Работы на строительных площадках»",
            "п.29 «Работы в металлургическом производстве»",
            "п.30 «Работы в горнодобывающей промышленности»",
            "п.31 «Работы в деревообрабатывающем производстве»",
            "п.32 «Работы в текстильной и швейной промышленности»",
            "п.33 «Профессии и работы»"
        ]

        for idx, factor in enumerate(harmful_factors, start=1):
            harmful_sheet.cell(row=idx, column=1, value=factor)

        # Диапазон справочника вредностей и именованный диапазон
        last_row = len(harmful_factors)
        harmful_range = f"Ref_Harm!$A$1:$A${last_row}"

        harmful_validation = DataValidation(
            type="list",
            # Для диапазона используем ссылку с '=' — так Excel надёжнее распознаёт источник списка
            formula1=f"={harmful_range}",
            allow_blank=True,
            showDropDown=False,  # False показывает стрелку! (особенность openpyxl)
            showInputMessage=True,
            showErrorMessage=True
        )
        harmful_validation.error = "❌ Пожалуйста, выберите профессиональную вредность из выпадающего списка"
        harmful_validation.errorTitle = "Неверное значение"
        harmful_validation.prompt = "📋 Кликните на стрелку справа и выберите категорию профессиональной вредности согласно приказу № ҚР ДСМ-131/2020\n\nДоступно 33 категории (п.1 - п.33)"
        harmful_validation.promptTitle = "Выбор профессиональной вредности"
        ws.add_data_validation(harmful_validation)

        # Колонка "Профессиональная вредность" (K)
        harmful_validation.add(f"K{data_start_row}:K{data_end_row}")

        # 3) Валидация для числовых полей (Общий стаж и Стаж по должности)
        # Колонка H - Общий стаж
        experience_validation = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        experience_validation.error = "❌ Введите только целое число (количество лет стажа)"
        experience_validation.errorTitle = "Неверный формат"
        experience_validation.prompt = "🔢 Введите количество лет общего стажа работы\n\nПримеры: 20, 15, 5\n\nВводите только цифры без текста!"
        experience_validation.promptTitle = "Общий стаж (лет)"
        ws.add_data_validation(experience_validation)
        experience_validation.add(f"H{data_start_row}:H{data_end_row}")

        # Колонка I - Стаж по занимаемой должности
        position_experience_validation = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        position_experience_validation.error = "❌ Введите только целое число (количество лет стажа)"
        position_experience_validation.errorTitle = "Неверный формат"
        position_experience_validation.prompt = "🔢 Введите количество лет стажа по текущей должности\n\nПримеры: 18, 10, 3\n\nВводите только цифры без текста!"
        position_experience_validation.promptTitle = "Стаж по должности (лет)"
        ws.add_data_validation(position_experience_validation)
        position_experience_validation.add(f"I{data_start_row}:I{data_end_row}")

        # 3.5) Валидация для телефона (колонка E)
        # Используем текстовую валидацию с проверкой длины
        phone_validation = DataValidation(
            type="textLength",
            operator="between",
            formula1="10",
            formula2="12",
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        phone_validation.error = "❌ Номер телефона должен содержать от 10 до 12 цифр\n\nПримеры правильных форматов:\n• 77001234567 (11 цифр)\n• 7001234567 (10 цифр)\n• +77001234567 (12 символов)"
        phone_validation.errorTitle = "Неверный формат телефона"
        phone_validation.prompt = "📱 Введите номер телефона (10-12 цифр)\n\nФормат:\n• 77001234567\n• +77001234567\n• 87001234567\n\n⚠️ Номер будет преобразован в формат 7XXXXXXXXXX для WhatsApp"
        phone_validation.promptTitle = "Телефон (необязательно)"
        ws.add_data_validation(phone_validation)
        phone_validation.add(f"E{data_start_row}:E{data_end_row}")
        
        # Устанавливаем формат ячеек телефона как текст, чтобы избежать научной нотации
        for row in range(data_start_row, data_end_row + 1):
            cell = ws[f'E{row}']
            cell.number_format = '@'  # @ означает текстовый формат

        # 4) Валидация для дат (Дата рождения и Дата последнего медосмотра)
        # Колонка C - Дата рождения
        birth_date_validation = DataValidation(
            type="custom",
            formula1='=AND(LEN(C5)=10, ISNUMBER(DATEVALUE(C5)))',
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        birth_date_validation.error = "❌ Введите дату в формате ДД.ММ.ГГГГ\n\nПример: 29.03.1976"
        birth_date_validation.errorTitle = "Неверный формат даты"
        birth_date_validation.prompt = "📅 Введите дату рождения в формате:\nДД.ММ.ГГГГ\n\nПримеры:\n• 29.03.1976\n• 15.05.1985\n• 01.01.1990"
        birth_date_validation.promptTitle = "Дата рождения"
        ws.add_data_validation(birth_date_validation)
        birth_date_validation.add(f"C{data_start_row}:C{data_end_row}")

        # Колонка J - Дата последнего медосмотра
        exam_date_validation = DataValidation(
            type="custom",
            formula1='=AND(LEN(J5)=10, ISNUMBER(DATEVALUE(J5)))',
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        exam_date_validation.error = "❌ Введите дату в формате ДД.ММ.ГГГГ\n\nПример: 22.01.2024"
        exam_date_validation.errorTitle = "Неверный формат даты"
        exam_date_validation.prompt = "📅 Введите дату последнего медосмотра в формате:\nДД.ММ.ГГГГ\n\nПримеры:\n• 22.01.2024\n• 15.03.2023\n• 01.12.2024"
        exam_date_validation.promptTitle = "Дата последнего медосмотра"
        ws.add_data_validation(exam_date_validation)
        exam_date_validation.add(f"J{data_start_row}:J{data_end_row}")

        # 5) Валидация для текстовых полей с подсказками
        # Колонка B - ФИО
        fio_validation = DataValidation(
            type="textLength",
            operator="greaterThan",
            formula1="0",
            allow_blank=False,
            showInputMessage=True,
            showErrorMessage=True
        )
        fio_validation.error = "❌ ФИО обязательно для заполнения"
        fio_validation.errorTitle = "Пустое поле"
        fio_validation.prompt = "👤 Введите полное ФИО сотрудника\n\nФормат: Фамилия Имя Отчество\n\nПримеры:\n• Иванов Иван Иванович\n• Петрова Мария Петровна"
        fio_validation.promptTitle = "ФИО сотрудника"
        ws.add_data_validation(fio_validation)
        fio_validation.add(f"B{data_start_row}:B{data_end_row}")

        # Колонка F - Объект или участок
        department_validation = DataValidation(
            type="textLength",
            operator="greaterThan",
            formula1="0",
            allow_blank=False,
            showInputMessage=True,
            showErrorMessage=True
        )
        department_validation.error = "❌ Объект или участок обязательны для заполнения"
        department_validation.errorTitle = "Пустое поле"
        department_validation.prompt = "🏢 Укажите место работы сотрудника\n\nПримеры:\n• ТОО \"Компания\" - Отдел продаж\n• Производственный участок №1\n• Административный корпус"
        department_validation.promptTitle = "Объект или участок"
        ws.add_data_validation(department_validation)
        department_validation.add(f"F{data_start_row}:F{data_end_row}")

        # Колонка G - Занимаемая должность
        position_validation = DataValidation(
            type="textLength",
            operator="greaterThan",
            formula1="0",
            allow_blank=False,
            showInputMessage=True,
            showErrorMessage=True
        )
        position_validation.error = "❌ Должность обязательна для заполнения"
        position_validation.errorTitle = "Пустое поле"
        position_validation.prompt = "💼 Укажите должность сотрудника\n\nПримеры:\n• Оператор станков с ЧПУ\n• Главный бухгалтер\n• Инженер-технолог\n• Водитель погрузчика"
        position_validation.promptTitle = "Занимаемая должность"
        ws.add_data_validation(position_validation)
        position_validation.add(f"G{data_start_row}:G{data_end_row}")

        # Колонка L - Примечание (необязательное, только подсказка)
        notes_validation = DataValidation(
            type="textLength",
            operator="lessThan",
            formula1="1000",
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=False
        )
        notes_validation.prompt = "📝 Дополнительная информация о сотруднике (необязательно)\n\nПримеры:\n• Работает по совместительству\n• Временно отсутствует\n• Находится в декретном отпуске"
        notes_validation.promptTitle = "Примечание"
        ws.add_data_validation(notes_validation)
        notes_validation.add(f"L{data_start_row}:L{data_end_row}")

        # Скрываем справочные листы для удобства пользователя
        gender_sheet.sheet_state = "hidden"
        harmful_sheet.sheet_state = "hidden"

        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=\"шаблон_список_контингента.xlsx\"'
        return response

    @action(detail=False, methods=['delete'])
    def delete_all(self, request):
        """Удаление всех сотрудников контингента для текущего пользователя"""
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            count = ContingentEmployee.objects.filter(user=user).count()
            ContingentEmployee.objects.filter(user=user).delete()
            return Response({'message': f'Удалено {count} сотрудников'}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)


class CalendarPlanViewSet(viewsets.ModelViewSet):
    serializer_class = CalendarPlanSerializer

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Экспорт календарного плана в PDF"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab library not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        try:
            plan = self.get_object()
            print(f"[PDF Export] Generating PDF for plan {plan.id}: {plan.department}")
            print(f"[PDF Export] Plan data - departments_info: {plan.departments_info}")
            
            # Регистрация шрифта для поддержки кириллицы
            try:
                # Пробуем найти и зарегистрировать шрифты с поддержкой кириллицы
                import os
                font_paths = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                    '/usr/share/fonts/dejavu/DejaVuSans.ttf',
                    '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
                    'DejaVuSans.ttf',  # Может быть в PATH
                ]
                
                font_found = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
                            # Пробуем найти Bold версию
                            bold_path = font_path.replace('DejaVuSans.ttf', 'DejaVuSans-Bold.ttf')
                            if os.path.exists(bold_path):
                                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_path))
                                font_name_bold = 'DejaVuSans-Bold'
                            else:
                                font_name_bold = 'DejaVuSans'  # Используем обычный шрифт вместо bold
                            
                            font_name = 'DejaVuSans'
                            font_found = True
                            print(f"[PDF Export] DejaVu fonts registered successfully from {font_path}")
                            break
                        except Exception as e:
                            print(f"[PDF Export] Failed to register font from {font_path}: {e}")
                            continue
                
                if not font_found:
                    print("[PDF Export] No suitable fonts found, using Helvetica (no Cyrillic support)")
                    font_name = 'Helvetica'
                    font_name_bold = 'Helvetica-Bold'
            except Exception as e:
                print(f"[PDF Export] Error during font registration: {e}")
                font_name = 'Helvetica'
                font_name_bold = 'Helvetica-Bold'
            
            # Создание PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Стиль для заголовка
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name_bold,
                fontSize=16,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=20,
                alignment=1  # Центрирование
            )
            
            # Стиль для подзаголовков
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontName=font_name_bold,
                fontSize=12,
                textColor=colors.HexColor('#374151'),
                spaceAfter=10
            )
            
            # Стиль для обычного текста
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                textColor=colors.HexColor('#1f2937')
            )
            
            # Заголовок
            elements.append(Paragraph('Календарный план медицинских осмотров', title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Основная информация
            # Создаем стиль для текста в ячейках с переносом
            cell_style = ParagraphStyle(
                'CellText',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#1f2937'),
                fontName=font_name,
                wordWrap='CJK',  # Перенос слов
                leading=12  # Межстрочный интервал
            )
            
            info_data = [
                ['Объект/участок:', Paragraph(plan.department, cell_style)],
                ['Период проведения:', f"{plan.start_date.strftime('%d.%m.%Y')} - {plan.end_date.strftime('%d.%m.%Y')}"],
                ['Количество сотрудников:', str(len(plan.employee_ids))],
                ['Статус:', self._get_status_text(plan.status)],
            ]
            
            if plan.contract:
                info_data.insert(0, ['Договор:', f"№{plan.contract.contract_number} от {plan.contract.contract_date.strftime('%d.%m.%Y')}"])
            
            info_table = Table(info_data, colWidths=[2.5*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), font_name_bold),
                ('FONTNAME', (1, 0), (1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Выравнивание по верху для многострочного текста
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Информация по участкам (если есть несколько)
            if plan.departments_info and len(plan.departments_info) > 1:
                elements.append(Paragraph('Участки и периоды проведения осмотров', subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                
                # Стиль для ячеек с переносом текста
                dept_cell_style = ParagraphStyle(
                    'DeptCellText',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#1f2937'),
                    fontName=font_name,
                    wordWrap='CJK',
                    leading=11
                )
                
                dept_data = [['№', 'Объект/участок', 'Период', 'Сотрудников']]
                for idx, dept_info in enumerate(plan.departments_info, 1):
                    # Поддержка как camelCase, так и snake_case ключей
                    department = dept_info.get('department', '')
                    start_date = dept_info.get('startDate') or dept_info.get('start_date', '')
                    end_date = dept_info.get('endDate') or dept_info.get('end_date', '')
                    employee_ids = dept_info.get('employeeIds') or dept_info.get('employee_ids', [])
                    
                    dept_data.append([
                        str(idx),
                        Paragraph(department, dept_cell_style),  # Используем Paragraph для переноса
                        f"{start_date} - {end_date}",
                        str(len(employee_ids))
                    ])
                
                dept_table = Table(dept_data, colWidths=[0.5*inch, 2.5*inch, 2*inch, 1.5*inch])
                dept_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Номер по центру
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Название участка слева
                    ('ALIGN', (2, 0), (-1, -1), 'CENTER'), # Остальное по центру
                    ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Выравнивание по верху
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ]))
                elements.append(dept_table)
                elements.append(Spacer(1, 0.3*inch))
            
            # Вредные факторы
            if plan.harmful_factors:
                elements.append(Paragraph('Вредные и опасные производственные факторы', subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                
                factors_text = ', '.join(plan.harmful_factors)
                elements.append(Paragraph(factors_text, normal_style))
                elements.append(Spacer(1, 0.3*inch))
            
            # Врачи
            if plan.selected_doctors:
                elements.append(Paragraph('Назначенные врачи', subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                
                try:
                    from .models import Doctor
                    doctors = Doctor.objects.filter(id__in=plan.selected_doctors)
                    if doctors.exists():
                        doctor_data = [['№', 'ФИО', 'Специализация', 'Кабинет']]
                        for idx, doctor in enumerate(doctors, 1):
                            doctor_data.append([
                                str(idx),
                                doctor.name,
                                doctor.specialization,
                                doctor.cabinet or '-'
                            ])
                        
                        doctor_table = Table(doctor_data, colWidths=[0.5*inch, 2.5*inch, 2*inch, 1.5*inch])
                        doctor_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                            ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('FONTNAME', (0, 1), (-1, -1), font_name),
                            ('FONTSIZE', (0, 1), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                            ('LEFTPADDING', (0, 0), (-1, -1), 8),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                        ]))
                        elements.append(doctor_table)
                        elements.append(Spacer(1, 0.3*inch))
                except Exception as e:
                    pass
            
            # Список сотрудников (контингент)
            if plan.employee_ids:
                elements.append(Paragraph('Список работников, подлежащих медицинскому осмотру', subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                
                try:
                    from .models import ContingentEmployee
                    
                    # Получаем сотрудников по ID
                    employees = ContingentEmployee.objects.filter(id__in=plan.employee_ids).order_by('department', 'name')
                    
                    if employees.exists():
                        # Стиль для ячеек таблицы сотрудников
                        emp_cell_style = ParagraphStyle(
                            'EmpCellText',
                            parent=styles['Normal'],
                            fontSize=8,
                            textColor=colors.HexColor('#1f2937'),
                            fontName=font_name,
                            wordWrap='CJK',
                            leading=10
                        )
                        
                        # Заголовок таблицы
                        emp_data = [['№', 'ФИО', 'Должность', 'Объект/участок', 'Вредные факторы']]
                        
                        # Добавляем сотрудников
                        for idx, emp in enumerate(employees, 1):
                            harmful_factors = ', '.join(emp.harmful_factors) if emp.harmful_factors else '-'
                            emp_data.append([
                                str(idx),
                                Paragraph(emp.name, emp_cell_style),
                                Paragraph(emp.position, emp_cell_style),
                                Paragraph(emp.department, emp_cell_style),
                                Paragraph(harmful_factors, emp_cell_style)
                            ])
                        
                        # Создаем таблицу с оптимальными размерами колонок
                        emp_table = Table(emp_data, colWidths=[0.4*inch, 1.8*inch, 1.5*inch, 1.8*inch, 1.5*inch])
                        emp_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Номер по центру
                            ('ALIGN', (1, 1), (-1, -1), 'LEFT'),   # Остальное слева
                            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('FONTNAME', (0, 1), (-1, -1), font_name),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('TOPPADDING', (0, 0), (-1, -1), 4),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                            ('LEFTPADDING', (0, 0), (-1, -1), 4),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                        ]))
                        elements.append(emp_table)
                        elements.append(Spacer(1, 0.3*inch))
                        
                        # Итоговая информация
                        summary_style = ParagraphStyle(
                            'Summary',
                            parent=styles['Normal'],
                            fontSize=10,
                            textColor=colors.HexColor('#374151'),
                            fontName=font_name_bold
                        )
                        elements.append(Paragraph(f'Всего работников: {employees.count()}', summary_style))
                        
                except Exception as e:
                    print(f"[PDF Export] Error loading employees: {str(e)}")
                    pass
            
            # Генерация PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            # Безопасное имя файла (только латиница, цифры и подчеркивания)
            safe_department = ''.join(c if c.isalnum() or c in (' ', '_') else '_' for c in plan.department)
            safe_department = safe_department.replace(' ', '_')
            filename = f"calendar_plan_{safe_department}_{plan.start_date.strftime('%Y%m%d')}.pdf"
            # Используем URL-кодирование для поддержки русских символов
            from urllib.parse import quote
            encoded_filename = quote(filename)
            response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
            print(f"[PDF Export] PDF generated successfully: {filename}")
            
            return response
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[PDF Export] Error generating PDF: {str(e)}")
            print(f"[PDF Export] Traceback:\n{error_trace}")
            return Response({'error': str(e), 'details': error_trace}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _get_status_text(self, status_code):
        """Получить текстовое описание статуса"""
        status_map = {
            'draft': 'Черновик',
            'pending_clinic': 'Ожидает утверждения клиникой',
            'pending_employer': 'Ожидает утверждения работодателем',
            'approved': 'Утвержден',
            'rejected': 'Отклонен работодателем',
            'sent_to_ses': 'Отправлен в СЭС'
        }
        return status_map.get(status_code, status_code)

    def get_queryset(self):
        # Для POST запросов (создание) возвращаем полный queryset
        if self.action == 'create':
            return CalendarPlan.objects.all()
            
        user_id = self.request.query_params.get('user_id')
        contract_id = self.request.query_params.get('contract_id')
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                
                # Если указан contract_id, фильтруем только по этому договору
                if contract_id:
                    import sys
                    print(f"DEBUG: Запрос календарных планов по договору {contract_id} для пользователя {user.username}", file=sys.stderr)
                    
                    try:
                        contract = Contract.objects.get(id=contract_id)
                        # Проверяем права доступа к договору
                        if user.role == 'clinic' and contract.clinic == user:
                            queryset = CalendarPlan.objects.filter(contract=contract).order_by('-created_at')
                        elif user.role == 'employer' and (contract.employer == user or contract.employer_bin == user.registration_data.get('bin')):
                            queryset = CalendarPlan.objects.filter(contract=contract).order_by('-created_at')
                        else:
                            queryset = CalendarPlan.objects.none()
                        
                        print(f"DEBUG: Найдено {queryset.count()} планов по договору {contract_id}", file=sys.stderr)
                        return queryset
                    except Contract.DoesNotExist:
                        print(f"DEBUG: Договор {contract_id} не найден", file=sys.stderr)
                        return CalendarPlan.objects.none()
                
                # Если пользователь - клиника, показываем планы, созданные этой клиникой
                if user.role == 'clinic':
                    return CalendarPlan.objects.filter(user=user).order_by('-created_at')
                # Если пользователь - работодатель, показываем планы по договорам, где он является стороной
                elif user.role == 'employer':
                    user_bin = user.registration_data.get('bin') or user.registration_data.get('inn') if user.registration_data else None
                    contracts = Contract.objects.filter(
                        Q(employer=user) | (Q(employer_bin=user_bin) if user_bin else Q()),
                        status='approved'
                    )
                    return CalendarPlan.objects.filter(contract__in=contracts).order_by('-created_at')
                else:
                    return CalendarPlan.objects.filter(user=user).order_by('-created_at')
            except User.DoesNotExist:
                return CalendarPlan.objects.none()
        return CalendarPlan.objects.all().order_by('-created_at')
    
    def update(self, request, *args, **kwargs):
        """Обновление календарного плана с автоматическим созданием маршрутных листов при согласовании"""
        instance = self.get_object()
        old_status = instance.status
        new_status = request.data.get('status', old_status)
        rejection_reason = request.data.get('rejection_reason', '')
        
        # Если статус меняется на 'rejected', устанавливаем дату отклонения и причину
        if new_status == 'rejected' and old_status != 'rejected':
            from django.utils import timezone
            request.data['rejected_by_employer_at'] = timezone.now()
            if not rejection_reason:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'rejection_reason': 'Необходимо указать причину отклонения'})
        
        # Если статус меняется с 'rejected' на другой, очищаем причину и дату отклонения
        if old_status == 'rejected' and new_status != 'rejected':
            request.data['rejection_reason'] = ''
            request.data['rejected_by_employer_at'] = None
        
        # Выполняем стандартное обновление
        response = super().update(request, *args, **kwargs)
        
        # Получаем обновленный объект из базы данных
        updated_instance = self.get_object()
        actual_new_status = updated_instance.status
        
        # Если статус изменился на 'approved', создаем маршрутные листы
        if old_status != 'approved' and actual_new_status == 'approved':
            try:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Creating route sheets for calendar plan {updated_instance.id} (status changed from {old_status} to {actual_new_status})")
                self._create_route_sheets_for_plan(updated_instance)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating route sheets for calendar plan {updated_instance.id}: {str(e)}", exc_info=True)
                # Не прерываем обновление плана, даже если создание маршрутных листов не удалось
        
        return response
    
    def partial_update(self, request, *args, **kwargs):
        """Частичное обновление календарного плана с автоматическим созданием маршрутных листов при согласовании"""
        instance = self.get_object()
        old_status = instance.status
        new_status = request.data.get('status', old_status)
        rejection_reason = request.data.get('rejection_reason', '')
        
        # Если статус меняется на 'rejected', устанавливаем дату отклонения и причину
        if new_status == 'rejected' and old_status != 'rejected':
            from django.utils import timezone
            request.data['rejected_by_employer_at'] = timezone.now()
            if not rejection_reason:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'rejection_reason': 'Необходимо указать причину отклонения'})
        
        # Если статус меняется с 'rejected' на другой, очищаем причину и дату отклонения
        if old_status == 'rejected' and new_status != 'rejected':
            request.data['rejection_reason'] = ''
            request.data['rejected_by_employer_at'] = None
        
        # Выполняем стандартное частичное обновление
        response = super().partial_update(request, *args, **kwargs)
        
        # Получаем обновленный объект из базы данных
        updated_instance = self.get_object()
        actual_new_status = updated_instance.status
        
        # Если статус изменился на 'approved', создаем маршрутные листы
        if old_status != 'approved' and actual_new_status == 'approved':
            try:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Creating route sheets for calendar plan {updated_instance.id} (status changed from {old_status} to {actual_new_status})")
                self._create_route_sheets_for_plan(updated_instance)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating route sheets for calendar plan {updated_instance.id}: {str(e)}", exc_info=True)
        
        return response
    
    def perform_update(self, serializer):
        """Обновление календарного плана - только клиника может редактировать свои планы"""
        user_id = self.request.data.get('user') or self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                plan = self.get_object()
                # Только клиника, создавшая план, может его редактировать
                if user.role == 'clinic' and plan.user == user:
                    # Можно редактировать только черновики
                    if plan.status != 'draft':
                        from rest_framework.exceptions import ValidationError
                        raise ValidationError({'status': 'Можно редактировать только черновики'})
                    serializer.save()
                else:
                    from rest_framework.exceptions import ValidationError
                    raise ValidationError({'user': 'У вас нет прав на редактирование этого плана'})
            except User.DoesNotExist:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'user': 'User not found'})
        else:
            serializer.save()
    
    def _create_route_sheets_for_plan(self, plan):
        """Автоматическое создание маршрутных листов для всех сотрудников календарного плана после согласования"""
        from datetime import datetime, timedelta
        import logging
        import json
        
        logger = logging.getLogger(__name__)
        logger.info(f"Starting route sheets creation for calendar plan {plan.id}")
        
        clinic_user = plan.user
        if not clinic_user or clinic_user.role != 'clinic':
            logger.warning(f"Plan {plan.id} user is not a clinic user: {clinic_user}")
            return
        
        logger.info(f"Clinic user: {clinic_user.id}, role: {clinic_user.role}")
        
        # Получаем список врачей из календарного плана
        selected_doctor_ids = plan.selected_doctors or []
        doctors_map = {}
        if selected_doctor_ids:
            doctors = Doctor.objects.filter(id__in=selected_doctor_ids, user=clinic_user)
            for doctor in doctors:
                doctors_map[doctor.specialization] = {
                    'id': str(doctor.id),
                    'name': doctor.name,
                    'cabinet': doctor.cabinet or '',
                    'specialization': doctor.specialization,
                }
        
        # Обрабатываем информацию по участкам
        departments_info = plan.departments_info or []
        if not departments_info:
            # Если нет информации по участкам, используем общие даты
            departments_info = [{
                'department': plan.department,
                'startDate': plan.start_date.isoformat() if hasattr(plan.start_date, 'isoformat') else str(plan.start_date),
                'endDate': plan.end_date.isoformat() if hasattr(plan.end_date, 'isoformat') else str(plan.end_date),
                'employeeIds': plan.employee_ids or [],
            }]
        
        logger.info(f"Processing {len(departments_info)} departments for plan {plan.id}")
        
        harmful_factors = plan.harmful_factors or []
        created_count = 0
        
        # Проверяем, что есть сотрудники для обработки
        total_employee_ids = set()
        for dept_info in departments_info:
            employee_ids = dept_info.get('employeeIds') or dept_info.get('employee_ids', [])
            total_employee_ids.update([str(eid) for eid in employee_ids])
        
        logger.info(f"Total unique employee IDs in plan: {len(total_employee_ids)}, IDs: {list(total_employee_ids)}")
        
        # Создаем маршрутные листы для каждого участка
        for dept_info in departments_info:
            department = dept_info.get('department', plan.department)
            start_date_str = dept_info.get('startDate') or dept_info.get('start_date', '')
            end_date_str = dept_info.get('endDate') or dept_info.get('end_date', '')
            employee_ids = dept_info.get('employeeIds') or dept_info.get('employee_ids', [])
            
            logger.info(f"Processing department {department}, employee_ids from plan: {employee_ids}, types: {[type(eid).__name__ for eid in employee_ids]}")
            
            # Парсим даты
            try:
                if isinstance(start_date_str, str):
                    start_date = datetime.strptime(start_date_str.split('T')[0], '%Y-%m-%d').date()
                else:
                    start_date = start_date_str
                
                if isinstance(end_date_str, str):
                    end_date = datetime.strptime(end_date_str.split('T')[0], '%Y-%m-%d').date()
                else:
                    end_date = end_date_str
            except:
                continue
            
            # Получаем сотрудников этого участка
            # Преобразуем employee_ids в строки и числа для поиска
            employee_ids_str = [str(eid) for eid in employee_ids]
            employee_ids_int = []
            for eid in employee_ids:
                try:
                    employee_ids_int.append(int(eid))
                except (ValueError, TypeError):
                    pass
            
            logger.info(f"Searching for employees with IDs: str={employee_ids_str}, int={employee_ids_int}")
            
            # Ищем сотрудников по ID независимо от роли пользователя
            # Сотрудники могут быть загружены как работодателем, так и клиникой по договору
            employees = ContingentEmployee.objects.filter(
                Q(id__in=employee_ids_str) | Q(id__in=employee_ids_int)
            )
            
            logger.info(f"Found {employees.count()} employees for department {department} (searching for {len(employee_ids)} IDs)")
            
            if employees.count() > 0:
                logger.info(f"Employee details: IDs={[str(e.id) for e in employees]}, Users={[e.user.id if e.user else 'no user' for e in employees]}, Roles={[e.user.role if e.user else 'no user' for e in employees]}")
            else:
                # Если не найдены, проверяем, существуют ли такие ID вообще
                all_employees_check = ContingentEmployee.objects.all()
                logger.warning(f"No employees found. Total employees in DB: {all_employees_check.count()}, Sample IDs: {[str(e.id) for e in all_employees_check[:5]]}")
            
            # Создаем маршрутный лист для каждого сотрудника на дату начала периода
            for employee in employees:
                # Проверяем, не существует ли уже маршрутный лист
                existing = RouteSheet.objects.filter(
                    user=clinic_user,
                    patient_id=str(employee.id),
                    visit_date=start_date
                ).first()
                
                if existing:
                    continue
                
                # Получаем вредные факторы сотрудника
                harmful_factors_emp = []
                if employee.harmful_factors:
                    if isinstance(employee.harmful_factors, str):
                        try:
                            harmful_factors_emp = json.loads(employee.harmful_factors)
                        except:
                            harmful_factors_emp = [f.strip() for f in employee.harmful_factors.split(',') if f.strip()]
                    elif isinstance(employee.harmful_factors, list):
                        harmful_factors_emp = employee.harmful_factors
                
                # Используем вредные факторы из календарного плана, если они указаны
                if harmful_factors:
                    harmful_factors_emp = list(set(harmful_factors_emp + harmful_factors))
                
                # Генерируем услуги для маршрутного листа
                route_sheet_viewset = RouteSheetViewSet()
                route_sheet_viewset.request = type('obj', (object,), {'data': {}, 'query_params': {}, 'user': clinic_user})()
                services = route_sheet_viewset._generate_services_for_position(
                    employee.position,
                    harmful_factors_emp,
                    user_id=clinic_user.id
                )
                
                # Назначаем врачей из календарного плана
                if doctors_map:
                    for service in services:
                        specialization = service.get('specialization', service.get('name', ''))
                        if specialization in doctors_map:
                            doctor_info = doctors_map[specialization]
                            service['doctorId'] = doctor_info['id']
                            service['cabinet'] = doctor_info['cabinet'] or service.get('cabinet', 'Не указан')
                
                # Создаем маршрутный лист
                route_sheet = RouteSheet.objects.create(
                    user=clinic_user,
                    patient_id=str(employee.id),
                    patient_name=employee.name,
                    iin=employee.iin or '',
                    position=employee.position,
                    department=employee.department,
                    visit_date=start_date,
                    services=services,
                )
                
                # Создаем лабораторные и функциональные исследования
                route_sheet_viewset._create_required_tests(route_sheet, employee.position, harmful_factors_emp)
                
                created_count += 1
                logger.info(f"Created route sheet ID {route_sheet.id} for employee {employee.id} ({employee.name}) on {start_date}, patient_id={route_sheet.patient_id}")
        
        total_employees_in_plan = sum(len(dept_info.get('employeeIds') or dept_info.get('employee_ids', [])) for dept_info in departments_info)
        logger.info(f"Successfully created {created_count} route sheets for calendar plan {plan.id} (total employees in plan: {total_employees_in_plan})")
    
    def perform_destroy(self, instance):
        """Удаление календарного плана - только клиника может удалять свои планы"""
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                # Только клиника, создавшая план, может его удалить
                if user.role == 'clinic' and instance.user == user:
                    # Можно удалять только черновики
                    if instance.status != 'draft':
                        from rest_framework.exceptions import ValidationError
                        raise ValidationError({'status': 'Можно удалять только черновики'})
                    instance.delete()
                else:
                    from rest_framework.exceptions import ValidationError
                    raise ValidationError({'user': 'У вас нет прав на удаление этого плана'})
            except User.DoesNotExist:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'user': 'User not found'})
        else:
            instance.delete()

    def perform_create(self, serializer):
        # Получаем user_id из данных запроса или из текущего пользователя
        user_id = self.request.data.get('user')
        contract_id = self.request.data.get('contract')
        
        print(f"[CalendarPlan] Creating calendar plan - user_id: {user_id}, contract_id: {contract_id}")
        print(f"[CalendarPlan] Request data: {self.request.data}")
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                print(f"[CalendarPlan] User found: {user.phone}, role: {user.role}")
                contract = None
                
                # Если указан договор, проверяем его
                if contract_id:
                    try:
                        contract = Contract.objects.get(id=contract_id)
                        print(f"[CalendarPlan] Contract found: {contract.contract_number}, status: {contract.status}")
                        
                        # Проверяем, что договор подтвержден или исполнен
                        if contract.status not in ['approved', 'executed']:
                            print(f"[CalendarPlan] Contract status invalid: {contract.status}")
                            raise ValidationError({'contract': f'Договор должен быть подтвержден или исполнен. Текущий статус: {contract.status}'})
                        
                        # Проверяем, что пользователь связан с договором
                        if user.role == 'clinic':
                            if contract.clinic != user:
                                print(f"[CalendarPlan] User is not clinic of this contract")
                                raise ValidationError({'contract': 'Вы не являетесь стороной этого договора'})
                        elif user.role == 'employer':
                            if contract.employer != user:
                                # Проверяем по БИН
                                reg_data = user.registration_data or {}
                                user_bin = reg_data.get('bin') or reg_data.get('inn')
                                print(f"[CalendarPlan] Checking BIN: user_bin={user_bin}, contract_bin={contract.employer_bin}")
                                if not user_bin or str(user_bin).strip() != str(contract.employer_bin or '').strip():
                                    raise ValidationError({'contract': 'Вы не являетесь стороной этого договора'})
                    except Contract.DoesNotExist:
                        print(f"[CalendarPlan] Contract not found: {contract_id}")
                        raise ValidationError({'contract': 'Договор не найден'})
                else:
                    # Если договор не указан, ищем подтвержденный договор
                    print(f"[CalendarPlan] No contract specified, searching for approved contract")
                    if user.role == 'clinic':
                        contract = Contract.objects.filter(clinic=user, status__in=['approved', 'executed']).first()
                    elif user.role == 'employer':
                        contract = Contract.objects.filter(
                            Q(employer=user) | Q(employer_bin=user.registration_data.get('bin', '') if user.registration_data else ''),
                            status__in=['approved', 'executed']
                        ).first()
                    
                    if not contract:
                        print(f"[CalendarPlan] No approved contract found for user")
                        raise ValidationError({'contract': 'Необходимо выбрать подтвержденный договор для создания календарного плана'})
                    print(f"[CalendarPlan] Auto-selected contract: {contract.contract_number}")
                
                print(f"[CalendarPlan] Saving calendar plan with contract: {contract.contract_number if contract else 'None'}")
                serializer.save(user=user, contract=contract)
                print(f"[CalendarPlan] Calendar plan created successfully")
            except User.DoesNotExist:
                print(f"[CalendarPlan] User not found: {user_id}")
                raise ValidationError({'user': 'User not found'})
            except Exception as e:
                print(f"[CalendarPlan] Unexpected error: {type(e).__name__}: {str(e)}")
                raise
        else:
            print(f"[CalendarPlan] No user_id provided, saving without user")
            serializer.save()


class RouteSheetViewSet(viewsets.ModelViewSet):
    serializer_class = RouteSheetSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                
                # Если пользователь - клиника, возвращаем маршрутные листы, созданные этой клиникой
                if user.role == 'clinic':
                    return RouteSheet.objects.filter(user=user).order_by('-created_at')
                # Если пользователь - работодатель, возвращаем маршрутные листы для его сотрудников
                elif user.role == 'employer':
                    try:
                        # Находим договоры с этим работодателем (используем те же статусы, что и для контингента)
                        user_bin = user.registration_data.get('bin') or user.registration_data.get('inn') if user.registration_data else None
                        allowed_statuses = ['approved', 'active', 'in_progress', 'sent', 'pending_approval']
                        contracts = Contract.objects.filter(
                            Q(employer=user) | (Q(employer_bin=user_bin) if user_bin else Q()),
                            status__in=allowed_statuses
                        )
                        
                        # Находим всех сотрудников работодателя (используем ту же логику, что и для контингента)
                        all_employees = ContingentEmployee.objects.filter(
                            Q(user=user) | Q(contract__in=contracts)
                        ).distinct()
                        
                        # Получаем ID сотрудников
                        employee_ids = [str(emp.id) for emp in all_employees]
                        
                        # Возвращаем маршрутные листы для этих сотрудников
                        if employee_ids:
                            return RouteSheet.objects.filter(patient_id__in=employee_ids).order_by('-created_at')
                        else:
                            return RouteSheet.objects.none()
                    except Exception as e:
                        logger.error(f"Error getting route sheets for employer {user.id}: {str(e)}")
                        return RouteSheet.objects.none()
                else:
                    return RouteSheet.objects.filter(user=user).order_by('-created_at')
            except User.DoesNotExist:
                return RouteSheet.objects.none()
            except Exception as e:
                logger.error(f"Error in RouteSheetViewSet.get_queryset: {str(e)}")
                return RouteSheet.objects.none()
        return RouteSheet.objects.all().order_by('-created_at')

    @action(detail=True, methods=['patch'])
    def update_service_status(self, request, pk=None):
        """Обновление статуса услуги в маршрутном листе. Только врач соответствующей специализации может отмечать услугу."""
        service_id = request.data.get('service_id')
        new_status = request.data.get('status')  # Переименовано, чтобы не конфликтовать с импортом
        
        if not service_id or new_status not in ['pending', 'completed']:
            return Response({'error': 'Invalid service_id or status'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            route_sheet = self.get_object()
            services = route_sheet.services if isinstance(route_sheet.services, list) else []
            
            # Находим услугу
            target_service = None
            for service in services:
                if str(service.get('id')) == str(service_id):
                    target_service = service
                    break
            
            if not target_service:
                return Response({'error': 'Service not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Проверка прав доступа
            # Получаем текущего пользователя
            user_id = request.query_params.get('user_id') or request.data.get('user_id')
            if not user_id:
                return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                current_user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Если пользователь - клиника (менеджер, регистратура), запрещаем редактирование
            if current_user.role == 'clinic' and current_user.clinic_role in ['manager', 'receptionist', None]:
                return Response({
                    'error': 'Только врач может отмечать услуги как выполненные. Клиника может только просматривать статус.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Если пользователь - врач, проверяем соответствие специализации
            if current_user.role == 'clinic' and current_user.clinic_role == 'doctor':
                # Получаем врача из базы данных
                # Врач связан с клиникой через поле user (это клиника, а не сам врач)
                # Нужно найти врача, который принадлежит этой клинике
                try:
                    # Ищем врача, который принадлежит клинике текущего пользователя
                    doctor = Doctor.objects.filter(user=current_user).first()
                    if not doctor:
                        return Response({
                            'error': 'Врач не найден в базе данных. Убедитесь, что ваш профиль врача создан.'
                        }, status=status.HTTP_404_NOT_FOUND)
                    
                    doctor_specialization = doctor.specialization
                    service_specialization = target_service.get('specialization') or target_service.get('name')
                    
                    # Проверяем соответствие специализации
                    if doctor_specialization != service_specialization:
                        return Response({
                            'error': f'Вы можете отмечать только услуги по вашей специализации ({doctor_specialization}). Данная услуга: {service_specialization}'
                        }, status=status.HTTP_403_FORBIDDEN)
                except Exception as e:
                    return Response({
                        'error': f'Ошибка при проверке прав доступа: {str(e)}'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Обновляем статус услуги
            target_service['status'] = new_status
            route_sheet.services = services
            route_sheet.save()
            
            serializer = RouteSheetSerializer(route_sheet)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc() if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def create_by_iin(self, request):
        phone = request.data.get('phone')  # Поиск по номеру телефона
        iin = request.data.get('iin')  # Оставляем для обратной совместимости
        name = request.data.get('name')  # Оставляем для обратной совместимости
        user_id = request.data.get('user_id')
        visit_date = request.data.get('visit_date')
        
        try:
            user = User.objects.get(id=user_id)
            employee = None
            
            # Приоритет поиска: сначала по телефону
            if phone:
                # Очищаем телефон от пробелов и форматируем
                phone_clean = ''.join(filter(str.isdigit, phone))
                employee = ContingentEmployee.objects.filter(user=user, phone__icontains=phone_clean).first()
            
            # Если не найден по телефону, ищем по ИИН (для обратной совместимости)
            if not employee and iin:
                employee = ContingentEmployee.objects.filter(user=user, iin=iin).first()
            
            # Если не найден по ИИН, ищем по ФИО (для обратной совместимости)
            if not employee and name:
                employee = ContingentEmployee.objects.filter(user=user, name__icontains=name).first()
            
            # Если все еще не найден, пробуем найти среди всех работодателей (для клиники)
            if not employee:
                if user.role == 'clinic':
                    if phone:
                        phone_clean = ''.join(filter(str.isdigit, phone))
                        employee = ContingentEmployee.objects.filter(phone__icontains=phone_clean, user__role='employer').first()
                    elif iin:
                        employee = ContingentEmployee.objects.filter(iin=iin, user__role='employer').first()
                    elif name:
                        employee = ContingentEmployee.objects.filter(name__icontains=name, user__role='employer').first()
                else:
                    return Response({'error': 'Сотрудник не найден в базе данных'}, status=status.HTTP_404_NOT_FOUND)
            
            if not employee:
                return Response({'error': 'Сотрудник не найден. Проверьте номер телефона.'}, status=status.HTTP_404_NOT_FOUND)
            
            # Проверяем, есть ли утвержденный календарный план для этого сотрудника
            from datetime import datetime
            visit_date_obj = datetime.strptime(visit_date, '%Y-%m-%d').date() if isinstance(visit_date, str) else visit_date
            
            calendar_plan = CalendarPlan.objects.filter(
                user=user if user.role == 'clinic' else employee.user,
                employee_ids__contains=[str(employee.id)],
                status__in=['approved', 'sent_to_ses'],
                start_date__lte=visit_date_obj,
                end_date__gte=visit_date_obj
            ).first()
            
            if not calendar_plan:
                return Response({
                    'error': 'Для этого сотрудника нет утвержденного календарного плана на указанную дату'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if route sheet already exists
            existing = RouteSheet.objects.filter(
                user=user,
                patient_id=str(employee.id),
                visit_date=visit_date_obj
            ).first()
            
            if existing:
                serializer = RouteSheetSerializer(existing)
                return Response(serializer.data, status=status.HTTP_200_OK)
            
            # Generate services based on position and harmful factors
            # Обрабатываем harmful_factors - может быть списком или строкой JSON
            harmful_factors_list = []
            if employee.harmful_factors:
                if isinstance(employee.harmful_factors, str):
                    import json
                    try:
                        harmful_factors_list = json.loads(employee.harmful_factors)
                    except:
                        # Если не JSON, разбиваем по запятой
                        harmful_factors_list = [f.strip() for f in employee.harmful_factors.split(',') if f.strip()]
                elif isinstance(employee.harmful_factors, list):
                    harmful_factors_list = employee.harmful_factors
            
            # Передаем user_id в метод генерации услуг через request
            services = self._generate_services_for_position(employee.position, harmful_factors_list, user_id=user.id)
            
            # Убеждаемся, что services не пустой
            if not services:
                # Если услуги не сгенерированы, добавляем базовые
                from datetime import datetime, timedelta
                start_time = datetime.strptime('09:00', '%H:%M')
                services = [{
                    'id': '0',
                    'name': 'Терапевт',
                    'cabinet': '11',
                    'doctorId': '1',
                    'specialization': 'Терапевт',
                    'time': start_time.strftime('%H:%M'),
                    'status': 'pending',
                }]
            
            route_sheet = RouteSheet.objects.create(
                user=user,
                patient_id=str(employee.id),
                patient_name=employee.name,
                iin=employee.iin or '',
                position=employee.position,
                department=employee.department,
                visit_date=visit_date_obj,
                services=services,
            )
            
            # Автоматически создаем лабораторные и функциональные исследования
            self._create_required_tests(route_sheet, employee.position, harmful_factors_list)
            
            serializer = RouteSheetSerializer(route_sheet)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            import traceback
            return Response({'error': str(e), 'traceback': traceback.format_exc() if settings.DEBUG else None}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def generate_qr_code(self, request, pk=None):
        """Генерация QR-кода для маршрутного листа"""
        if not QRCODE_AVAILABLE:
            return Response({'error': 'QRCode library not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        try:
            route_sheet = self.get_object()
            
            # Формируем данные для QR-кода
            qr_data = {
                'route_sheet_id': str(route_sheet.id),
                'patient_id': route_sheet.patient_id,
                'patient_name': route_sheet.patient_name,
                'iin': route_sheet.iin,
                'visit_date': route_sheet.visit_date.isoformat() if route_sheet.visit_date else None,
                'url': f'{settings.FRONTEND_URL if hasattr(settings, "FRONTEND_URL") else "http://localhost:3000"}/dashboard/clinic/route-sheet?id={route_sheet.id}',
            }
            
            # Создаем QR-код
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(json.dumps(qr_data, ensure_ascii=False))
            qr.make(fit=True)
            
            # Создаем изображение
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Сохраняем в буфер
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='image/png')
            response['Content-Disposition'] = f'inline; filename="qr_route_sheet_{route_sheet.id}.png"'
            return response
            
        except RouteSheet.DoesNotExist:
            return Response({'error': 'Route sheet not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _generate_services_for_position(self, position, harmful_factors, user_id=None):
        """
        Генерация услуг на основе должности и вредных факторов.
        Использует реальных врачей из базы данных клиники.
        Вредные факторы имеют приоритет - если указаны, добавляются дополнительные услуги.
        """
        from datetime import datetime, timedelta
        
        # Получаем user_id из параметра или из запроса
        if not user_id:
            if hasattr(self, 'request') and self.request:
                user_id = self.request.data.get('user_id') or self.request.query_params.get('user_id')
            
            # Если user_id не найден, пытаемся получить из контекста запроса
            if not user_id and hasattr(self, 'request') and self.request.user.is_authenticated:
                # Для клиники используем текущего пользователя
                if self.request.user.role == 'clinic':
                    user_id = self.request.user.id
        
        # Получаем врачей клиники из базы данных
        doctors_by_specialization = {}
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                if user.role == 'clinic':
                    doctors = Doctor.objects.filter(user=user)
                    for doctor in doctors:
                        spec = doctor.specialization
                        if spec not in doctors_by_specialization:
                            doctors_by_specialization[spec] = []
                        doctors_by_specialization[spec].append({
                            'id': str(doctor.id),
                            'name': doctor.name,
                            'cabinet': doctor.cabinet or '',
                            'specialization': doctor.specialization,
                        })
            except User.DoesNotExist:
                pass
        
        def get_doctor_for_specialization(specialization):
            """Получить первого доступного врача по специализации"""
            if specialization in doctors_by_specialization and doctors_by_specialization[specialization]:
                doctor = doctors_by_specialization[specialization][0]
                return {
                    'name': doctor['specialization'],
                    'cabinet': doctor['cabinet'] or 'Не указан',
                    'doctorId': doctor['id'],
                    'specialization': doctor['specialization'],
                }
            return None
        
        # Базовые услуги для всех должностей
        base_services = []
        therapist = get_doctor_for_specialization('Терапевт')
        if therapist:
            base_services.append(therapist)
        else:
            base_services.append({'name': 'Терапевт', 'cabinet': 'Не указан', 'doctorId': '', 'specialization': 'Терапевт'})
        
        # Услуги на основе должности
        position_to_specializations = {
            'Бухгалтер': ['Терапевт', 'Окулист', 'Невропатолог'],
            'Сварщик': ['Профпатолог', 'ЛОР', 'Окулист', 'Хирург', 'Невропатолог', 'Терапевт', 'Рентгенолог'],
            'Водитель': ['Профпатолог', 'Окулист', 'Невропатолог', 'Терапевт'],
        }
        
        # Начинаем с базовых услуг для должности
        required_specializations = position_to_specializations.get(position, ['Терапевт'])
        services = []
        for spec in required_specializations:
            doctor = get_doctor_for_specialization(spec)
            if doctor:
                services.append(doctor)
            else:
                # Если врача нет в БД, используем заглушку
                services.append({
                    'name': spec,
                    'cabinet': 'Не указан',
                    'doctorId': '',
                    'specialization': spec,
                })
        
        # Добавляем услуги на основе вредных факторов
        harmful_factors_lower = [str(f).lower() for f in harmful_factors] if harmful_factors else []
        
        # Маппинг вредных факторов на дополнительные специализации
        factor_to_specializations = {
            'шум': ['ЛОР'],
            'вибрация': ['Невропатолог'],
            'пыль': ['ЛОР', 'Рентгенолог'],
            'химические вещества': ['Терапевт', 'Профпатолог'],
            'излучение': ['Окулист', 'Рентгенолог'],
            'высотные работы': ['Невропатолог', 'Окулист'],
        }
        
        # Добавляем услуги на основе вредных факторов
        added_services = []
        for factor in harmful_factors_lower:
            for key, specializations in factor_to_specializations.items():
                if key in factor:
                    for spec in specializations:
                        # Проверяем, нет ли уже такой услуги
                        if not any(s['specialization'] == spec for s in services):
                            doctor = get_doctor_for_specialization(spec)
                            if doctor:
                                added_services.append(doctor)
                            else:
                                added_services.append({
                                    'name': spec,
                                    'cabinet': 'Не указан',
                                    'doctorId': '',
                                    'specialization': spec,
                                })
        
        services.extend(added_services)
        
        # Удаляем дубликаты по специализации
        seen = set()
        unique_services = []
        for service in services:
            if service['specialization'] not in seen:
                seen.add(service['specialization'])
                unique_services.append(service)
        services = unique_services
        
        # Генерируем время для каждого кабинета (начиная с 9:00, по 15 минут на кабинет)
        start_time = datetime.strptime('09:00', '%H:%M')
        result_services = []
        for idx, service in enumerate(services):
            service_time = start_time + timedelta(minutes=idx * 15)
            result_services.append({
                'id': str(idx),
                'name': service['name'],
                'cabinet': service['cabinet'],
                'doctorId': service['doctorId'],
                'specialization': service.get('specialization', service['name']),
                'time': service_time.strftime('%H:%M'),
                'status': 'pending',
            })
        
        return result_services

    def _create_required_tests(self, route_sheet, position, harmful_factors):
        """Автоматическое создание лабораторных и функциональных исследований на основе должности и вредных факторов"""
        # Базовые лабораторные исследования для всех
        base_lab_tests = [
            {'test_type': 'Общий анализ крови', 'test_name': 'ОАК'},
            {'test_type': 'Общий анализ мочи', 'test_name': 'ОАМ'},
        ]
        
        # Лабораторные исследования на основе должности
        position_lab_tests = {
            'Сварщик': [
                {'test_type': 'Биохимия крови', 'test_name': 'Биохимия'},
                {'test_type': 'Анализ на тяжелые металлы', 'test_name': 'Тяжелые металлы'},
            ],
            'Водитель': [
                {'test_type': 'Анализ на алкоголь и наркотики', 'test_name': 'Алкоголь/наркотики'},
            ],
        }
        
        # Функциональные исследования на основе должности
        position_func_tests = {
            'Сварщик': [
                {'test_type': 'Рентген легких', 'test_name': 'Флюорография'},
            ],
            'Водитель': [
                {'test_type': 'ЭКГ', 'test_name': 'Электрокардиограмма'},
            ],
        }
        
        # Добавляем лабораторные исследования
        lab_tests = base_lab_tests.copy()
        if position in position_lab_tests:
            lab_tests.extend(position_lab_tests[position])
        
        # Добавляем на основе вредных факторов
        harmful_factors_lower = [str(f).lower() for f in harmful_factors] if harmful_factors else []
        if any('пыль' in f for f in harmful_factors_lower):
            lab_tests.append({'test_type': 'Анализ мокроты', 'test_name': 'Мокрота'})
        
        # Создаем лабораторные исследования
        for test_data in lab_tests:
            LaboratoryTest.objects.create(
                route_sheet=route_sheet,
                patient_id=route_sheet.patient_id,
                patient_name=route_sheet.patient_name,
                test_type=test_data['test_type'],
                test_name=test_data['test_name'],
                status='pending',
            )
        
        # Добавляем функциональные исследования
        func_tests = []
        if position in position_func_tests:
            func_tests.extend(position_func_tests[position])
        
        # Добавляем на основе вредных факторов
        if any('шум' in f for f in harmful_factors_lower):
            func_tests.append({'test_type': 'Аудиометрия', 'test_name': 'Проверка слуха'})
        if any('вибрация' in f for f in harmful_factors_lower):
            func_tests.append({'test_type': 'Спирометрия', 'test_name': 'Функция дыхания'})
        
        # Создаем функциональные исследования
        for test_data in func_tests:
            FunctionalTest.objects.create(
                route_sheet=route_sheet,
                patient_id=route_sheet.patient_id,
                patient_name=route_sheet.patient_name,
                test_type=test_data['test_type'],
                test_name=test_data['test_name'],
                status='pending',
            )


class DoctorExaminationViewSet(viewsets.ModelViewSet):
    serializer_class = DoctorExaminationSerializer

    def get_queryset(self):
        patient_id = self.request.query_params.get('patient_id')
        if patient_id:
            return DoctorExamination.objects.filter(patient_id=patient_id).order_by('-examination_date')
        return DoctorExamination.objects.all().order_by('-examination_date')

    @action(detail=False, methods=['get'])
    def patient_history(self, request):
        """Получение истории осмотров пациента с группировкой по датам"""
        patient_id = request.query_params.get('patient_id')
        iin = request.query_params.get('iin')
        
        if not patient_id and not iin:
            return Response({'error': 'Необходимо указать patient_id или iin'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Если передан iin, находим patient_id
            if iin:
                employee = ContingentEmployee.objects.filter(iin=iin).first()
                if employee:
                    patient_id = str(employee.id)
                else:
                    return Response({'error': 'Пациент не найден'}, status=status.HTTP_404_NOT_FOUND)
            
            # Получаем все осмотры пациента
            examinations = DoctorExamination.objects.filter(patient_id=patient_id).order_by('-examination_date')
            
            # Группируем по датам
            history_by_date = {}
            for exam in examinations:
                exam_date = exam.examination_date.strftime('%Y-%m-%d') if exam.examination_date else 'Без даты'
                if exam_date not in history_by_date:
                    history_by_date[exam_date] = {
                        'date': exam_date,
                        'examinations': [],
                        'route_sheets': [],
                    }
                
                # Добавляем осмотр
                exam_data = {
                    'id': exam.id,
                    'doctor_name': exam.doctor_name,
                    'specialization': exam.specialization,
                    'conclusion': exam.conclusion,
                    'notes': exam.notes,
                    'recommendations': exam.recommendations,
                    'examination_date': exam.examination_date.isoformat() if exam.examination_date else None,
                }
                history_by_date[exam_date]['examinations'].append(exam_data)
            
            # Получаем маршрутные листы
            route_sheets = RouteSheet.objects.filter(patient_id=patient_id).order_by('-visit_date')
            for rs in route_sheets:
                rs_date = rs.visit_date.strftime('%Y-%m-%d') if rs.visit_date else 'Без даты'
                if rs_date not in history_by_date:
                    history_by_date[rs_date] = {
                        'date': rs_date,
                        'examinations': [],
                        'route_sheets': [],
                    }
                
                # Подсчитываем завершенные услуги
                completed_services = sum(1 for s in (rs.services or []) if s.get('status') == 'completed')
                total_services = len(rs.services or [])
                
                route_sheet_data = {
                    'id': rs.id,
                    'visit_date': rs.visit_date.isoformat() if rs.visit_date else None,
                    'completed_services': completed_services,
                    'total_services': total_services,
                    'services': rs.services or [],
                }
                history_by_date[rs_date]['route_sheets'].append(route_sheet_data)
            
            # Получаем экспертизы
            expertises = Expertise.objects.filter(patient_id=patient_id).order_by('-created_at')
            for exp in expertises:
                exp_date = exp.created_at.strftime('%Y-%m-%d') if exp.created_at else 'Без даты'
                if exp_date not in history_by_date:
                    history_by_date[exp_date] = {
                        'date': exp_date,
                        'examinations': [],
                        'route_sheets': [],
                        'expertises': [],
                    }
                
                if 'expertises' not in history_by_date[exp_date]:
                    history_by_date[exp_date]['expertises'] = []
                
                expertise_data = {
                    'id': exp.id,
                    'final_verdict': exp.final_verdict,
                    'health_group': exp.health_group,
                    'created_at': exp.created_at.isoformat() if exp.created_at else None,
                }
                history_by_date[exp_date]['expertises'].append(expertise_data)
            
            # Преобразуем в список и сортируем по дате (новые сначала)
            history_list = list(history_by_date.values())
            history_list.sort(key=lambda x: x['date'], reverse=True)
            
            return Response({
                'patient_id': patient_id,
                'history': history_list,
                'total_visits': len(history_list),
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExpertiseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpertiseSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                return Expertise.objects.filter(user=user)
            except User.DoesNotExist:
                return Expertise.objects.none()
        return Expertise.objects.all()

    def _check_examination_completeness(self, patient_id, route_sheet_id=None):
        """Проверка полноты осмотра перед экспертизой"""
        errors = []
        
        # Проверяем наличие маршрутного листа
        if route_sheet_id:
            try:
                route_sheet = RouteSheet.objects.get(id=route_sheet_id)
            except RouteSheet.DoesNotExist:
                errors.append('Маршрутный лист не найден')
                return errors
        else:
            route_sheet = RouteSheet.objects.filter(patient_id=patient_id).first()
            if not route_sheet:
                errors.append('Маршрутный лист не найден')
                return errors
        
        # Проверяем, что все услуги завершены
        services = route_sheet.services if isinstance(route_sheet.services, list) else []
        pending_services = [s for s in services if s.get('status') != 'completed']
        if pending_services:
            errors.append(f'Не все врачи завершили осмотр. Осталось: {len(pending_services)}')
        
        # Проверяем лабораторные исследования
        lab_tests = LaboratoryTest.objects.filter(patient_id=patient_id, route_sheet=route_sheet)
        pending_lab = lab_tests.filter(status__in=['pending', 'in_progress'])
        if pending_lab.exists():
            errors.append(f'Не все лабораторные исследования завершены. Осталось: {pending_lab.count()}')
        
        # Проверяем функциональные исследования
        func_tests = FunctionalTest.objects.filter(patient_id=patient_id, route_sheet=route_sheet)
        pending_func = func_tests.filter(status__in=['pending', 'in_progress'])
        if pending_func.exists():
            errors.append(f'Не все функциональные исследования завершены. Осталось: {pending_func.count()}')
        
        return errors

    @action(detail=False, methods=['get'])
    def final_act_stats(self, request):
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False)
            if department:
                queryset = queryset.filter(department=department)
            
            # Calculate occupational diseases
            occupational_diseases = 0
            for exp in queryset:
                if exp.doctor_conclusions:
                    for conclusion in exp.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('notes', '').lower().find('профзаболевание') != -1:
                            occupational_diseases += 1
                            break
            
            stats = {
                'totalExamined': queryset.count(),
                'healthy': queryset.filter(final_verdict='fit').count(),
                'temporaryContraindications': queryset.filter(final_verdict='temporary_unfit').count(),
                'permanentContraindications': queryset.filter(final_verdict='permanent_unfit').count(),
                'occupationalDiseases': occupational_diseases,
            }
            
            return Response(stats, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def health_plan_items(self, request):
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False).exclude(final_verdict='fit')
            if department:
                queryset = queryset.filter(department=department)
            
            items = []
            for expertise in queryset:
                # Get recommendations from doctor conclusions
                recommendations = []
                if expertise.doctor_conclusions:
                    for conclusion in expertise.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('recommendations'):
                            recommendations.append(conclusion['recommendations'])
                        elif isinstance(conclusion, dict) and conclusion.get('notes') and conclusion.get('conclusion') == 'unhealthy':
                            recommendations.append(conclusion['notes'])
                
                recommendation_text = '; '.join(recommendations) if recommendations else (expertise.reason or 'Требуется дополнительное обследование')
                items.append({
                    'patientId': expertise.patient_id,
                    'employeeName': expertise.patient_name,
                    'position': expertise.position,
                    'recommendation': recommendation_text,
                })
            
            return Response(items, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    def _update_contract_status(self, expertise):
        """Обновление статуса договора при завершении медосмотра"""
        # Находим сотрудника по ИИН
        employee = ContingentEmployee.objects.filter(iin=expertise.iin).first()
        if not employee or not employee.user:
            return
        
        # Находим активные договоры работодателя с клиникой
        contracts = Contract.objects.filter(
            employer=employee.user,
            clinic=expertise.user,
            status__in=['in_progress', 'partially_executed']
        )
        
        for contract in contracts:
            # Подсчитываем завершенные медосмотры
            total_people = contract.people_count
            if total_people == 0:
                continue
            
            # Получаем все экспертизы для сотрудников работодателя
            expertises = Expertise.objects.filter(
                user=contract.clinic,
                final_verdict__isnull=False
            )
            
            # Подсчитываем уникальных пациентов с завершенными экспертизами
            completed_patients = set()
            for exp in expertises:
                emp = ContingentEmployee.objects.filter(
                    user=contract.employer,
                    iin=exp.iin
                ).first()
                if emp:
                    completed_patients.add(exp.patient_id)
            
            completed_count = len(completed_patients)
            percentage = (completed_count / total_people) * 100 if total_people > 0 else 0
            
            # Обновляем статус договора
            old_status = contract.status
            if percentage == 100:
                contract.status = 'executed'
            elif percentage > 0:
                contract.status = 'partially_executed'
            
            if old_status != contract.status:
                contract.save()
    
    def _assign_health_group(self, expertise):
        """Автоматическое присвоение группы здоровья на основе вердикта и заключений врачей"""
        if not expertise.final_verdict:
            return None
        
        # Если годен и нет патологий - группа 1
        if expertise.final_verdict == 'fit':
            has_pathology = False
            if expertise.doctor_conclusions:
                for conclusion in expertise.doctor_conclusions:
                    if isinstance(conclusion, dict) and conclusion.get('conclusion') == 'unhealthy':
                        has_pathology = True
                        break
            
            if not has_pathology:
                return '1'  # Здоровый
            else:
                return '2'  # Практически здоровый (есть незначительные отклонения)
        
        # Если временная непригодность - группа 3 или 4
        elif expertise.final_verdict == 'temporary_unfit':
            # Если есть признаки воздействия ВПФ - группа 3
            if expertise.doctor_conclusions:
                for conclusion in expertise.doctor_conclusions:
                    if isinstance(conclusion, dict):
                        notes = conclusion.get('notes', '').lower()
                        if 'впф' in notes or 'вредный фактор' in notes:
                            return '3'  # Имеет признаки воздействия ВПФ
            
            return '4'  # Требует динамического наблюдения
        
        # Если постоянная непригодность - группа 5 или 6
        elif expertise.final_verdict == 'permanent_unfit':
            # Если требуется реабилитация - группа 6
            if expertise.reason and ('реабилитац' in expertise.reason.lower() or 'профпатолог' in expertise.reason.lower()):
                return '6'  # Требует реабилитации/профпатологии
            
            return '5'  # Требует лечения
        
        return None

    def perform_create(self, serializer):
        """Создание экспертизы с проверкой полноты осмотра"""
        patient_id = self.request.data.get('patient_id')
        route_sheet_id = self.request.data.get('route_sheet_id')
        
        # Проверяем полноту осмотра
        errors = self._check_examination_completeness(patient_id, route_sheet_id)
        if errors:
            raise ValidationError({'error': '; '.join(errors)})
        
        serializer.save()

    def perform_update(self, serializer):
        instance = serializer.save()
        # Автоматически присваиваем группу здоровья при обновлении вердикта
        if instance.final_verdict and not instance.health_group:
            health_group = self._assign_health_group(instance)
            if health_group:
                instance.health_group = health_group
                # Если группа 4, 5 или 6 - требуется направление
                if health_group in ['4', '5', '6']:
                    instance.requires_referral = True
                    if health_group == '6':
                        instance.referral_type = 'rehabilitation'
                    elif health_group == '5':
                        instance.referral_type = 'specialist'
                    else:
                        instance.referral_type = 'profpathology'
                instance.save()
        
        # Автоматически создаем направление при вынесении вердикта, если требуется
        if instance.final_verdict and instance.requires_referral and instance.referral_type and not instance.referral_sent:
            # Проверяем, не создано ли уже направление
            existing_referral = Referral.objects.filter(
                patient_id=instance.patient_id,
                referral_type=instance.referral_type,
                status__in=['created', 'sent', 'accepted', 'in_progress']
            ).first()
            
            if not existing_referral:
                # Создаем направление
                referral = Referral.objects.create(
                    user=instance.user,
                    expertise=instance,
                    patient_id=instance.patient_id,
                    patient_name=instance.patient_name,
                    iin=instance.iin,
                    referral_type=instance.referral_type,
                    reason=instance.reason or f'Требуется направление согласно группе здоровья {instance.health_group}',
                    status='created',
                )
                instance.referral_sent = True
                instance.referral_date = timezone.now().date()
                instance.save()
        
        # Обновляем статус договора при завершении экспертизы
        if instance.final_verdict:
            self._update_contract_status(instance)
        
        serializer.instance = instance

    @action(detail=False, methods=['get'])
    def check_readiness(self, request):
        """Проверка готовности пациента к экспертизе"""
        patient_id = request.query_params.get('patient_id')
        route_sheet_id = request.query_params.get('route_sheet_id')
        
        if not patient_id:
            return Response({'error': 'patient_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        errors = self._check_examination_completeness(patient_id, route_sheet_id)
        is_ready = len(errors) == 0
        
        return Response({
            'is_ready': is_ready,
            'errors': errors,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def export_summary_report_pdf(self, request):
        """Экспорт сводного отчета в PDF"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab library not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False)
            
            if department:
                queryset = queryset.filter(department=department)
            if start_date:
                queryset = queryset.filter(verdict_date__gte=start_date)
            if end_date:
                queryset = queryset.filter(verdict_date__lte=end_date)
            
            # Подсчет статистики
            total = queryset.count()
            healthy = queryset.filter(final_verdict='fit').count()
            temporary = queryset.filter(final_verdict='temporary_unfit').count()
            permanent = queryset.filter(final_verdict='permanent_unfit').count()
            
            # Подсчет профзаболеваний
            occupational_diseases = 0
            for exp in queryset:
                if exp.doctor_conclusions:
                    for conclusion in exp.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('notes', '').lower().find('профзаболевание') != -1:
                            occupational_diseases += 1
                            break
            
            # Создание PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Заголовок
            elements.append(Paragraph('СВОДНЫЙ ОТЧЕТ', title_style))
            elements.append(Paragraph('по результатам медицинских осмотров', styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Дата формирования
            elements.append(Paragraph(f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")}', styles['Normal']))
            if department:
                elements.append(Paragraph(f'Отдел: {department}', styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Общая статистика
            data = [
                ['Показатель', 'Количество'],
                ['Всего осмотрено', str(total)],
                ['Здоровы', str(healthy)],
                ['Временные противопоказания', str(temporary)],
                ['Постоянные противопоказания', str(permanent)],
                ['Выявлено профзаболеваний', str(occupational_diseases)],
            ]
            
            table = Table(data, colWidths=[4*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Статистика по отделам
            department_stats = {}
            for exp in queryset:
                dept = exp.department
                if dept not in department_stats:
                    department_stats[dept] = {'total': 0, 'healthy': 0, 'temporary': 0, 'permanent': 0}
                department_stats[dept]['total'] += 1
                if exp.final_verdict == 'fit':
                    department_stats[dept]['healthy'] += 1
                elif exp.final_verdict == 'temporary_unfit':
                    department_stats[dept]['temporary'] += 1
                elif exp.final_verdict == 'permanent_unfit':
                    department_stats[dept]['permanent'] += 1
            
            if department_stats:
                elements.append(Paragraph('Статистика по отделам', styles['Heading2']))
                dept_data = [['Отдел', 'Всего', 'Здоровы', 'Временные', 'Постоянные']]
                for dept, stats in department_stats.items():
                    dept_data.append([dept, str(stats['total']), str(stats['healthy']), str(stats['temporary']), str(stats['permanent'])])
                
                dept_table = Table(dept_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
                dept_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(dept_table)
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="summary_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_summary_report_excel(self, request):
        """Экспорт сводного отчета в Excel"""
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False)
            
            if department:
                queryset = queryset.filter(department=department)
            if start_date:
                queryset = queryset.filter(verdict_date__gte=start_date)
            if end_date:
                queryset = queryset.filter(verdict_date__lte=end_date)
            
            # Подсчет статистики
            total = queryset.count()
            healthy = queryset.filter(final_verdict='fit').count()
            temporary = queryset.filter(final_verdict='temporary_unfit').count()
            permanent = queryset.filter(final_verdict='permanent_unfit').count()
            
            # Подсчет профзаболеваний
            occupational_diseases = 0
            for exp in queryset:
                if exp.doctor_conclusions:
                    for conclusion in exp.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('notes', '').lower().find('профзаболевание') != -1:
                            occupational_diseases += 1
                            break
            
            # Создание Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Сводный отчет"
            
            # Заголовок
            ws['A1'] = 'СВОДНЫЙ ОТЧЕТ'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = 'по результатам медицинских осмотров'
            ws['A3'] = f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            if department:
                ws['A4'] = f'Отдел: {department}'
            
            # Общая статистика
            row = 6
            ws[f'A{row}'] = 'Показатель'
            ws[f'B{row}'] = 'Количество'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Всего осмотрено'
            ws[f'B{row}'] = total
            row += 1
            ws[f'A{row}'] = 'Здоровы'
            ws[f'B{row}'] = healthy
            row += 1
            ws[f'A{row}'] = 'Временные противопоказания'
            ws[f'B{row}'] = temporary
            row += 1
            ws[f'A{row}'] = 'Постоянные противопоказания'
            ws[f'B{row}'] = permanent
            row += 1
            ws[f'A{row}'] = 'Выявлено профзаболеваний'
            ws[f'B{row}'] = occupational_diseases
            
            # Статистика по отделам
            row += 2
            ws[f'A{row}'] = 'Статистика по отделам'
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            ws[f'A{row}'] = 'Отдел'
            ws[f'B{row}'] = 'Всего'
            ws[f'C{row}'] = 'Здоровы'
            ws[f'D{row}'] = 'Временные'
            ws[f'E{row}'] = 'Постоянные'
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = Font(bold=True)
            
            department_stats = {}
            for exp in queryset:
                dept = exp.department
                if dept not in department_stats:
                    department_stats[dept] = {'total': 0, 'healthy': 0, 'temporary': 0, 'permanent': 0}
                department_stats[dept]['total'] += 1
                if exp.final_verdict == 'fit':
                    department_stats[dept]['healthy'] += 1
                elif exp.final_verdict == 'temporary_unfit':
                    department_stats[dept]['temporary'] += 1
                elif exp.final_verdict == 'permanent_unfit':
                    department_stats[dept]['permanent'] += 1
            
            for dept, stats in department_stats.items():
                row += 1
                ws[f'A{row}'] = dept
                ws[f'B{row}'] = stats['total']
                ws[f'C{row}'] = stats['healthy']
                ws[f'D{row}'] = stats['temporary']
                ws[f'E{row}'] = stats['permanent']
            
            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="summary_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_final_act_pdf(self, request):
        """Экспорт заключительного акта в PDF"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab library not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False)
            if department:
                queryset = queryset.filter(department=department)
            
            # Подсчет статистики
            total = queryset.count()
            healthy = queryset.filter(final_verdict='fit').count()
            temporary = queryset.filter(final_verdict='temporary_unfit').count()
            permanent = queryset.filter(final_verdict='permanent_unfit').count()
            
            # Подсчет профзаболеваний
            occupational_diseases = 0
            for exp in queryset:
                if exp.doctor_conclusions:
                    for conclusion in exp.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('notes', '').lower().find('профзаболевание') != -1:
                            occupational_diseases += 1
                            break
            
            # Создание PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Заголовок
            elements.append(Paragraph('ЗАКЛЮЧИТЕЛЬНЫЙ АКТ', title_style))
            elements.append(Paragraph('по результатам обязательных медицинских осмотров', styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Дата формирования
            elements.append(Paragraph(f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")}', styles['Normal']))
            if department:
                elements.append(Paragraph(f'Отдел: {department}', styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Статистика
            data = [
                ['Показатель', 'Количество'],
                ['Осмотрено', str(total)],
                ['Здоровы', str(healthy)],
                ['Временные противопоказания', str(temporary)],
                ['Постоянные противопоказания', str(permanent)],
                ['Выявлено профзаболеваний', str(occupational_diseases)],
            ]
            
            table = Table(data, colWidths=[4*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.5*inch))
            
            # Подписи
            elements.append(Paragraph('Подписи:', styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph('Главный врач клиники: _________________', styles['Normal']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph('Представитель работодателя: _________________', styles['Normal']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph('Представитель СЭС: _________________', styles['Normal']))
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="final_act_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_final_act_excel(self, request):
        """Экспорт заключительного акта в Excel"""
        user_id = request.query_params.get('user_id')
        department = request.query_params.get('department')
        
        try:
            user = User.objects.get(id=user_id)
            queryset = Expertise.objects.filter(user=user, final_verdict__isnull=False)
            if department:
                queryset = queryset.filter(department=department)
            
            # Подсчет статистики
            total = queryset.count()
            healthy = queryset.filter(final_verdict='fit').count()
            temporary = queryset.filter(final_verdict='temporary_unfit').count()
            permanent = queryset.filter(final_verdict='permanent_unfit').count()
            
            # Подсчет профзаболеваний
            occupational_diseases = 0
            for exp in queryset:
                if exp.doctor_conclusions:
                    for conclusion in exp.doctor_conclusions:
                        if isinstance(conclusion, dict) and conclusion.get('notes', '').lower().find('профзаболевание') != -1:
                            occupational_diseases += 1
                            break
            
            # Создание Excel файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Заключительный акт"
            
            # Заголовок
            ws.merge_cells('A1:B1')
            title_cell = ws['A1']
            title_cell.value = 'ЗАКЛЮЧИТЕЛЬНЫЙ АКТ'
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:B2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = 'по результатам обязательных медицинских осмотров'
            subtitle_cell.font = Font(size=12)
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Дата формирования
            ws['A3'] = f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            if department:
                ws['A4'] = f'Отдел: {department}'
            
            # Статистика
            start_row = 6
            ws[f'A{start_row}'] = 'Показатель'
            ws[f'B{start_row}'] = 'Количество'
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            ws[f'A{start_row}'].fill = header_fill
            ws[f'A{start_row}'].font = header_font
            ws[f'B{start_row}'].fill = header_fill
            ws[f'B{start_row}'].font = header_font
            
            data_rows = [
                ['Осмотрено', total],
                ['Здоровы', healthy],
                ['Временные противопоказания', temporary],
                ['Постоянные противопоказания', permanent],
                ['Выявлено профзаболеваний', occupational_diseases],
            ]
            
            for idx, row_data in enumerate(data_rows, start=start_row + 1):
                ws[f'A{idx}'] = row_data[0]
                ws[f'B{idx}'] = row_data[1]
            
            # Подписи
            signature_row = start_row + len(data_rows) + 3
            ws[f'A{signature_row}'] = 'Подписи:'
            ws[f'A{signature_row}'].font = Font(bold=True)
            
            signatures = [
                'Главный врач клиники: _________________',
                'Представитель работодателя: _________________',
                'Представитель СЭС: _________________',
            ]
            
            for idx, signature in enumerate(signatures, start=signature_row + 1):
                ws[f'A{idx}'] = signature
            
            # Автоподбор ширины столбцов
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 20
            
            # Сохранение в буфер
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="final_act_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmergencyNotificationViewSet(viewsets.ModelViewSet):
    serializer_class = EmergencyNotificationSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                return EmergencyNotification.objects.filter(user=user)
            except User.DoesNotExist:
                return EmergencyNotification.objects.none()
        return EmergencyNotification.objects.all()

    @action(detail=True, methods=['post'])
    def send_notification(self, request, pk=None):
        """Отправка экстренного извещения в ТСБ/СЭБН и работодателю"""
        notification = self.get_object()
        
        try:
            # Симуляция отправки в ТСБ/СЭБН
            # В реальности здесь будет интеграция с API ТСБ/СЭБН
            notification.sent_to_tsb = True
            
            # Отправка работодателю
            notification.sent_to_employer = True
            
            notification.sent_at = timezone.now()
            notification.save()
            
            return Response({'message': 'Экстренное извещение отправлено в ТСБ/СЭБН и работодателю'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HealthImprovementPlanViewSet(viewsets.ModelViewSet):
    serializer_class = HealthImprovementPlanSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                return HealthImprovementPlan.objects.filter(user=user)
            except User.DoesNotExist:
                return HealthImprovementPlan.objects.none()
        return HealthImprovementPlan.objects.all()


class RecommendationTrackingViewSet(viewsets.ModelViewSet):
    serializer_class = RecommendationTrackingSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                return RecommendationTracking.objects.filter(user=user)
            except User.DoesNotExist:
                return RecommendationTracking.objects.none()
        return RecommendationTracking.objects.all()

    def perform_create(self, serializer):
        user_id = self.request.data.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                serializer.save(user=user)
            except User.DoesNotExist:
                raise ValidationError({'user': 'User not found'})
        else:
            serializer.save()

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        status = request.data.get('status')
        if status not in ['pending', 'in_progress', 'completed', 'cancelled']:
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)
        
        tracking = self.get_object()
        tracking.status = status
        tracking.notes = request.data.get('notes', tracking.notes)
        tracking.save()
        
        serializer = RecommendationTrackingSerializer(tracking)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DoctorViewSet(viewsets.ModelViewSet):
    serializer_class = DoctorSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                return Doctor.objects.filter(user=user)
            except User.DoesNotExist:
                return Doctor.objects.none()
        return Doctor.objects.all()

    def perform_create(self, serializer):
        user_id = self.request.data.get('user_id')
        if not user_id:
            raise ValidationError({'user_id': 'user_id is required'})
        
        try:
            clinic_user = User.objects.get(id=user_id)
            # Проверяем, что это клиника
            if clinic_user.role != 'clinic':
                raise ValidationError({'user': 'Only clinics can add doctors'})
            
            # Сохраняем врача
            doctor = serializer.save(user=clinic_user)
            
            # Создаем или получаем User для врача, если указан телефон
            doctor_user = None
            if doctor.phone:
                # Нормализуем телефон: убираем все символы, оставляем только цифры
                phone_cleaned = doctor.phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                # Если начинается с 8, заменяем на 7
                if phone_cleaned.startswith('8'):
                    phone_cleaned = '7' + phone_cleaned[1:]
                # Если не начинается с 7, добавляем 7
                if not phone_cleaned.startswith('7'):
                    phone_cleaned = '7' + phone_cleaned
                # Ограничиваем до 11 цифр
                phone_cleaned = phone_cleaned[:11]
                
                # Форматируем в стандартный формат: +7 (XXX) XXX-XX-XX
                if len(phone_cleaned) == 11:
                    phone_formatted = f"+{phone_cleaned[0]} ({phone_cleaned[1:4]}) {phone_cleaned[4:7]}-{phone_cleaned[7:9]}-{phone_cleaned[9:11]}"
                    
                    try:
                        doctor_user = User.objects.get(phone=phone_formatted)
                        # Обновляем роль, если нужно
                        if doctor_user.role != 'clinic' or doctor_user.clinic_role != 'doctor':
                            doctor_user.role = 'clinic'
                            doctor_user.clinic_role = 'doctor'
                            doctor_user.registration_data = {
                                'name': doctor.name,
                                'specialization': doctor.specialization,
                                'doctor_id': doctor.id,
                                'clinic_id': clinic_user.id,
                            }
                            doctor_user.registration_completed = True
                            doctor_user.save()
                    except User.DoesNotExist:
                        # Создаем нового пользователя для врача
                        doctor_user = User.objects.create_user(
                            username=phone_formatted,
                            phone=phone_formatted,
                            role='clinic',
                            clinic_role='doctor',
                            registration_completed=True,
                            registration_data={
                                'name': doctor.name,
                                'specialization': doctor.specialization,
                                'doctor_id': doctor.id,
                                'clinic_id': clinic_user.id,
                            }
                        )
            
            # Отправляем уведомление врачу
            if doctor_user and doctor.phone:
                self._send_doctor_notification(doctor, clinic_user, doctor_user.phone)
                
        except User.DoesNotExist:
            raise ValidationError({'user': 'User not found'})
    
    def _send_doctor_notification(self, doctor, clinic_user, phone):
        """Отправка уведомления врачу о добавлении в платформу через WhatsApp"""
        try:
            from django.conf import settings
            import requests
            
            # Форматируем телефон
            formatted_phone = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            if not formatted_phone.startswith('7'):
                formatted_phone = '7' + formatted_phone
            
            # Формируем сообщение
            clinic_name = clinic_user.registration_data.get('name', 'Клиника')
            frontend_url = settings.FRONTEND_URL if hasattr(settings, 'FRONTEND_URL') else 'http://localhost:3000'
            
            message = f"""Добрый день, {doctor.name}!

Вас добавили в платформу MedCRM от {clinic_name}.

Ваша специализация: {doctor.specialization}
{f'Кабинет: {doctor.cabinet}' if doctor.cabinet else ''}

Теперь вы можете войти в личный кабинет врача, используя свой номер телефона:
{phone}

Для входа перейдите по ссылке:
{frontend_url}/auth

Войдите используя номер телефона {phone} и код из WhatsApp.

Добро пожаловать в платформу!"""
            
            # Отправляем через Green API
            url = f"{settings.GREEN_API_URL}/waInstance{settings.GREEN_API_ID_INSTANCE}/sendMessage/{settings.GREEN_API_TOKEN}"
            payload = {
                "chatId": f"{formatted_phone}@c.us",
                "message": message
            }
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code != 200:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to send doctor notification: {response.status_code}")
        except Exception as e:
            # Логируем ошибку, но не прерываем создание врача
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send doctor notification: {str(e)}")


class LaboratoryTestViewSet(viewsets.ModelViewSet):
    serializer_class = LaboratoryTestSerializer

    def get_queryset(self):
        patient_id = self.request.query_params.get('patient_id')
        route_sheet_id = self.request.query_params.get('route_sheet_id')
        
        queryset = LaboratoryTest.objects.all()
        
        if patient_id:
            queryset = queryset.filter(patient_id=patient_id)
        if route_sheet_id:
            queryset = queryset.filter(route_sheet_id=route_sheet_id)
        
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        route_sheet_id = self.request.data.get('route_sheet_id')
        if route_sheet_id:
            try:
                route_sheet = RouteSheet.objects.get(id=route_sheet_id)
                serializer.save(route_sheet=route_sheet)
            except RouteSheet.DoesNotExist:
                raise ValidationError({'route_sheet_id': 'Route sheet not found'})
        else:
            serializer.save()


class FunctionalTestViewSet(viewsets.ModelViewSet):
    serializer_class = FunctionalTestSerializer

    def get_queryset(self):
        patient_id = self.request.query_params.get('patient_id')
        route_sheet_id = self.request.query_params.get('route_sheet_id')
        
        queryset = FunctionalTest.objects.all()
        
        if patient_id:
            queryset = queryset.filter(patient_id=patient_id)
        if route_sheet_id:
            queryset = queryset.filter(route_sheet_id=route_sheet_id)
        
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        route_sheet_id = self.request.data.get('route_sheet_id')
        if route_sheet_id:
            try:
                route_sheet = RouteSheet.objects.get(id=route_sheet_id)
                serializer.save(route_sheet=route_sheet)
            except RouteSheet.DoesNotExist:
                raise ValidationError({'route_sheet_id': 'Route sheet not found'})
        else:
            serializer.save()


class ReferralViewSet(viewsets.ModelViewSet):
    serializer_class = ReferralSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        patient_id = self.request.query_params.get('patient_id')
        status_filter = self.request.query_params.get('status')
        
        queryset = Referral.objects.all()
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                queryset = queryset.filter(user=user)
            except User.DoesNotExist:
                return Referral.objects.none()
        
        if patient_id:
            queryset = queryset.filter(patient_id=patient_id)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        user_id = self.request.data.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                serializer.save(user=user)
            except User.DoesNotExist:
                raise ValidationError({'user': 'User not found'})
        else:
            serializer.save()

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        """Обновление статуса направления"""
        new_status = request.data.get('status')
        if new_status not in ['created', 'sent', 'accepted', 'in_progress', 'completed', 'cancelled']:
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)
        
        referral = self.get_object()
        referral.status = new_status
        
        if new_status == 'sent':
            referral.sent_at = timezone.now()
        elif new_status == 'accepted':
            referral.accepted_at = timezone.now()
        elif new_status == 'completed':
            referral.completed_at = timezone.now()
        
        referral.save()
        serializer = ReferralSerializer(referral)
        return Response(serializer.data, status=status.HTTP_200_OK)


class PatientQueueViewSet(viewsets.ModelViewSet):
    serializer_class = PatientQueueSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        doctor_id = self.request.query_params.get('doctor_id')
        status_filter = self.request.query_params.get('status')
        date = self.request.query_params.get('date')
        
        queryset = PatientQueue.objects.all()
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                queryset = queryset.filter(user=user)
            except User.DoesNotExist:
                return PatientQueue.objects.none()
        
        if doctor_id:
            queryset = queryset.filter(doctor_id=doctor_id)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if date:
            queryset = queryset.filter(added_at__date=date)
        
        return queryset.order_by('priority', 'queue_number', 'added_at')

    def perform_create(self, serializer):
        user_id = self.request.data.get('user_id')
        if not user_id:
            raise ValidationError({'user_id': 'user_id is required'})
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            raise ValidationError({'user': 'User not found'})
        
        today = timezone.now().date()
        last_queue = PatientQueue.objects.filter(
            user=user,
            added_at__date=today
        ).order_by('-queue_number').first()
        
        queue_number = (last_queue.queue_number + 1) if last_queue else 1
        serializer.save(user=user, queue_number=queue_number)

    @action(detail=True, methods=['post'])
    def call_patient(self, request, pk=None):
        """Вызвать пациента к врачу"""
        queue_entry = self.get_object()
        if queue_entry.status != 'waiting':
            return Response({'error': 'Пациент уже вызван или находится на приеме'}, status=status.HTTP_400_BAD_REQUEST)
        
        queue_entry.status = 'called'
        queue_entry.called_at = timezone.now()
        queue_entry.save()
        
        serializer = self.get_serializer(queue_entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def start_examination(self, request, pk=None):
        """Начать прием пациента"""
        queue_entry = self.get_object()
        if queue_entry.status not in ['called', 'waiting']:
            return Response({'error': 'Неверный статус для начала приема'}, status=status.HTTP_400_BAD_REQUEST)
        
        queue_entry.status = 'in_progress'
        queue_entry.started_at = timezone.now()
        if not queue_entry.called_at:
            queue_entry.called_at = timezone.now()
        queue_entry.save()
        
        serializer = self.get_serializer(queue_entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def complete_examination(self, request, pk=None):
        """Завершить прием пациента"""
        queue_entry = self.get_object()
        if queue_entry.status != 'in_progress':
            return Response({'error': 'Пациент не находится на приеме'}, status=status.HTTP_400_BAD_REQUEST)
        
        queue_entry.status = 'completed'
        queue_entry.completed_at = timezone.now()
        queue_entry.save()
        
        serializer = self.get_serializer(queue_entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def skip_patient(self, request, pk=None):
        """Пропустить пациента"""
        queue_entry = self.get_object()
        if queue_entry.status == 'completed':
            return Response({'error': 'Нельзя пропустить завершенного пациента'}, status=status.HTTP_400_BAD_REQUEST)
        
        queue_entry.status = 'skipped'
        queue_entry.save()
        
        serializer = self.get_serializer(queue_entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def current_queue(self, request):
        """Получить текущую очередь"""
        user_id = request.query_params.get('user_id')
        doctor_id = request.query_params.get('doctor_id')
        
        queryset = self.get_queryset().filter(
            status__in=['waiting', 'called', 'in_progress']
        )
        
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                queryset = queryset.filter(user=user)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if doctor_id:
            queryset = queryset.filter(doctor_id=doctor_id)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def add_from_route_sheet(self, request):
        """Добавить пациента в очередь из маршрутного листа"""
        route_sheet_id = request.data.get('route_sheet_id')
        service_id = request.data.get('service_id')
        user_id = request.data.get('user_id')
        priority = request.data.get('priority', 'normal')
        
        if not route_sheet_id or not service_id or not user_id:
            return Response({'error': 'route_sheet_id, service_id and user_id are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            route_sheet = RouteSheet.objects.get(id=route_sheet_id)
            user = User.objects.get(id=user_id)
            
            services = route_sheet.services if isinstance(route_sheet.services, list) else []
            service = None
            for s in services:
                if str(s.get('id')) == str(service_id):
                    service = s
                    break
            
            if not service:
                return Response({'error': 'Service not found in route sheet'}, status=status.HTTP_404_NOT_FOUND)
            
            existing = PatientQueue.objects.filter(
                user=user,
                route_sheet=route_sheet,
                service_name=service.get('name'),
                status__in=['waiting', 'called', 'in_progress']
            ).first()
            
            if existing:
                return Response({'error': 'Пациент уже в очереди для этой услуги'}, status=status.HTTP_400_BAD_REQUEST)
            
            today = timezone.now().date()
            last_queue = PatientQueue.objects.filter(
                user=user,
                added_at__date=today
            ).order_by('-queue_number').first()
            
            queue_number = (last_queue.queue_number + 1) if last_queue else 1
            
            doctor_id = service.get('doctorId') or service.get('doctor_id')
            doctor = None
            if doctor_id:
                try:
                    doctor = Doctor.objects.get(id=doctor_id)
                except Doctor.DoesNotExist:
                    pass
            
            queue_entry = PatientQueue.objects.create(
                user=user,
                route_sheet=route_sheet,
                doctor=doctor,
                patient_id=route_sheet.patient_id,
                patient_name=route_sheet.patient_name,
                iin=route_sheet.iin,
                service_name=service.get('name'),
                cabinet=service.get('cabinet', ''),
                priority=priority,
                queue_number=queue_number,
            )
            
            serializer = self.get_serializer(queue_entry)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except RouteSheet.DoesNotExist:
            return Response({'error': 'Route sheet not found'}, status=status.HTTP_404_NOT_FOUND)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ContractViewSet(viewsets.ModelViewSet):
    """ViewSet для управления договорами между работодателем и клиникой"""
    serializer_class = ContractSerializer
    
    def get_serializer_context(self):
        """Передаем request в контекст сериализатора для проверки прав доступа"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                # Работодатель видит свои договоры
                if user.role == 'employer':
                    # Ищем договоры где employer = user ИЛИ employer_bin совпадает с БИН работодателя
                    reg_data = user.registration_data or {}
                    employer_bin = reg_data.get('bin') or reg_data.get('inn')
                    
                    if employer_bin:
                        # Нормализуем БИН для поиска
                        bin_normalized = ''.join(str(employer_bin).strip().split())
                        
                        # Ищем договоры по разным вариантам БИН
                        # 1. employer = user (прямая связь)
                        # 2. employer_bin совпадает с нормализованным БИН
                        # 3. employer_bin совпадает с исходным БИН (на случай если в БД хранится с пробелами)
                        contracts = Contract.objects.filter(
                            Q(employer=user) | 
                            Q(employer_bin=bin_normalized) | 
                            Q(employer_bin=employer_bin) |
                            Q(employer_bin=str(employer_bin).strip())
                        ).distinct()
                        
                        print(f"[ContractViewSet] Employer {user.id} BIN: {employer_bin}, normalized: {bin_normalized}")
                        print(f"[ContractViewSet] Found {contracts.count()} contracts")
                        
                        return contracts
                    else:
                        return Contract.objects.filter(employer=user)
                # Клиника видит свои договоры
                elif user.role == 'clinic':
                    # Клиника видит:
                    # 1. Договоры где она является основной клиникой (clinic=user)
                    # 2. Договоры переданные ей на субподряд (subcontractor_clinic=user)
                    # 3. Договоры которые она передала на субподряд (original_clinic=user)
                    return Contract.objects.filter(
                        Q(clinic=user) | Q(subcontractor_clinic=user) | Q(original_clinic=user)
                    )
                else:
                    return Contract.objects.none()
            except User.DoesNotExist:
                return Contract.objects.none()
        return Contract.objects.all()
    
    def create(self, request, *args, **kwargs):
        """Переопределяем create для установки clinic до валидации"""
        user_id = request.data.get('user_id')
        employer_bin = request.data.get('employer_bin')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            employer = None
            
            # Создаем копию данных для изменения
            data = dict(request.data.items()) if hasattr(request.data, 'items') else dict(request.data)
            
            # Определяем employer и clinic
            if user.role == 'clinic':
                # Если указан БИН, ищем работодателя
                if employer_bin:
                    # Нормализуем БИН
                    bin_normalized = ''.join(str(employer_bin).strip().split())
                    
                    print(f"[ContractCreate] Searching for employer with BIN: {employer_bin}, normalized: {bin_normalized}")
                    
                    # Ищем по полям 'bin' и 'inn' в registration_data
                    employers = User.objects.filter(role='employer')
                    print(f"[ContractCreate] Total employers in DB: {employers.count()}")
                    
                    for emp in employers:
                        reg_data = emp.registration_data or {}
                        emp_bin = str(reg_data.get('bin', '')).strip()
                        emp_inn = str(reg_data.get('inn', '')).strip()
                        
                        # Нормализуем для сравнения
                        emp_bin_normalized = ''.join(emp_bin.split())
                        emp_inn_normalized = ''.join(emp_inn.split())
                        
                        if emp_bin_normalized == bin_normalized or emp_inn_normalized == bin_normalized:
                            employer = emp
                            print(f"[ContractCreate] Found employer: {emp.id}, name: {reg_data.get('name', 'N/A')}")
                            break
                    
                    if not employer:
                        print(f"[ContractCreate] WARNING: Employer not found for BIN: {employer_bin}")
                        # Сохраняем нормализованный БИН для будущего связывания
                        data['employer_bin'] = bin_normalized
                
                # Добавляем clinic в данные для сериализатора
                data['clinic'] = user.id
                if employer:
                    data['employer'] = employer.id
                    print(f"[ContractCreate] Contract will be linked to employer: {employer.id}")
                else:
                    print(f"[ContractCreate] Contract will be created without employer link, only with employer_bin: {employer_bin}")
                # Устанавливаем is_subcontracted по умолчанию
                if 'is_subcontracted' not in data:
                    data['is_subcontracted'] = False
            elif user.role == 'employer':
                clinic_id = request.data.get('clinic_id')
                if not clinic_id:
                    return Response({'error': 'clinic_id is required for employer'}, status=status.HTTP_400_BAD_REQUEST)
                data['employer'] = user.id
                data['clinic'] = clinic_id
                # Устанавливаем is_subcontracted по умолчанию
                if 'is_subcontracted' not in data:
                    data['is_subcontracted'] = False
            else:
                return Response({'error': 'User must be either employer or clinic'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Создаем сериализатор с обновленными данными
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except ValidationError as e:
            return Response({'error': str(e), 'details': e.detail if hasattr(e, 'detail') else None}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc() if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def perform_create(self, serializer):
        """Сохраняем договор и отправляем уведомление"""
        employer_bin = self.request.data.get('employer_bin', '')
        employer_phone = self.request.data.get('employer_phone', '')
        user_id = self.request.data.get('user_id')
        
        # Нормализуем БИН для единообразного хранения
        normalized_bin = ''.join(str(employer_bin).strip().split()) if employer_bin else ''
        
        # Сохраняем договор (clinic и employer уже установлены в create)
        contract = serializer.save(
            employer_bin=normalized_bin, 
            employer_phone=employer_phone or ''
        )
        
        # Создаем запись в истории
        try:
            user = User.objects.get(id=user_id)
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='created',
                user=user,
                user_role=user.role,
                user_name=user_name,
                new_status=contract.status,
                comment=f'Договор создан клиникой'
            )
        except User.DoesNotExist:
            pass
        
        # Отправляем уведомление работодателю, если указан телефон
        if user_id and employer_phone:
            try:
                user = User.objects.get(id=user_id)
                if user.role == 'clinic' and contract.status == 'draft':
                    self._send_contract_notification(contract, employer_phone, employer_bin or '')
            except User.DoesNotExist:
                pass
    
    def _send_contract_notification(self, contract, phone, bin_number):
        """Отправка уведомления работодателю о договоре через WhatsApp"""
        try:
            from django.conf import settings
            import requests
            
            # Форматируем телефон
            formatted_phone = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            if not formatted_phone.startswith('7'):
                formatted_phone = '7' + formatted_phone
            
            # Формируем сообщение
            clinic_name = contract.clinic.registration_data.get('name', 'Клиника')
            message = f"""Добрый день!

Вам отправлен договор на согласование от {clinic_name}.

Номер договора: {contract.contract_number}
Дата: {contract.contract_date}
Сумма: {contract.amount} тенге
Количество сотрудников: {contract.people_count}

Для подписания договора перейдите по ссылке:
{settings.FRONTEND_URL if hasattr(settings, 'FRONTEND_URL') else 'http://localhost:3000'}/dashboard/employer/contracts?bin={bin_number}

Если у вас нет аккаунта, зарегистрируйтесь по этой ссылке."""
            
            # Отправляем через Green API
            url = f"{settings.GREEN_API_URL}/waInstance{settings.GREEN_API_ID_INSTANCE}/sendMessage/{settings.GREEN_API_TOKEN}"
            payload = {
                "chatId": f"{formatted_phone}@c.us",
                "message": message
            }
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code == 200:
                # Обновляем статус договора на "отправлен"
                contract.status = 'sent'
                contract.sent_at = timezone.now()
                contract.save()
            else:
                # Даже если отправка не удалась, меняем статус на "отправлен" для отображения
                contract.status = 'sent'
                contract.sent_at = timezone.now()
                contract.save()
        except Exception as e:
            # Логируем ошибку, но не прерываем создание договора
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send contract notification: {str(e)}")
    
    def _send_subcontract_notification(self, contract, phone, bin_number, subcontract_amount, is_registered):
        """Отправка уведомления клинике-субподрядчику о передаче договора на субподряд через WhatsApp"""
        try:
            from django.conf import settings
            import requests
            
            # Форматируем телефон
            formatted_phone = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            if not formatted_phone.startswith('7'):
                formatted_phone = '7' + formatted_phone
            
            # Формируем сообщение
            original_clinic_name = contract.original_clinic.registration_data.get('name', 'Клиника') if contract.original_clinic and contract.original_clinic.registration_data else 'Клиника'
            frontend_url = settings.FRONTEND_URL if hasattr(settings, 'FRONTEND_URL') else 'http://localhost:3000'
            
            if is_registered:
                message = f"""Добрый день!

Вам передан договор на субподряд от {original_clinic_name}.

Номер договора: {contract.contract_number}
Дата договора: {contract.contract_date}
Количество сотрудников: {contract.people_count}
Сумма субподряда: {subcontract_amount} тенге

Для просмотра и принятия договора перейдите по ссылке:
{frontend_url}/dashboard/clinic/contracts

Войдите в систему используя свой номер телефона."""
            else:
                message = f"""Добрый день!

Вам передан договор на субподряд от {original_clinic_name}.

Номер договора: {contract.contract_number}
Дата договора: {contract.contract_date}
Количество сотрудников: {contract.people_count}
Сумма субподряда: {subcontract_amount} тенге
{f'БИН: {bin_number}' if bin_number else ''}

Для работы с договором зарегистрируйтесь в системе:
{frontend_url}/auth

После регистрации вы сможете принять или отклонить договор на субподряд."""
            
            # Отправляем через Green API
            url = f"{settings.GREEN_API_URL}/waInstance{settings.GREEN_API_ID_INSTANCE}/sendMessage/{settings.GREEN_API_TOKEN}"
            payload = {
                "chatId": f"{formatted_phone}@c.us",
                "message": message
            }
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code == 200:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Subcontract notification sent to {formatted_phone}")
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send subcontract notification: {response.status_code} - {response.text}")
        except Exception as e:
            # Логируем ошибку, но не прерываем процесс
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send subcontract notification: {str(e)}")
    
    @action(detail=True, methods=['post'])
    def subcontract(self, request, pk=None):
        """Передача договора на субподряд другой клинике"""
        try:
            contract = self.get_object()
            subcontractor_clinic_id = request.data.get('subcontractor_clinic_id')
            subcontractor_clinic_bin = request.data.get('subcontractor_clinic_bin')
            subcontractor_clinic_phone = request.data.get('subcontractor_clinic_phone')
            subcontract_amount = request.data.get('subcontract_amount')
            user_id = request.data.get('user_id')
            
            if not user_id:
                return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if not subcontract_amount:
                return Response({'error': 'subcontract_amount is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что пользователь является владельцем договора (оригинальной клиникой)
            user = User.objects.get(id=user_id)
            if user.role != 'clinic':
                return Response({'error': 'Only clinic can subcontract this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            # Определяем оригинальную клинику: если есть original_clinic, то она, иначе clinic
            original_clinic = contract.original_clinic if contract.original_clinic else contract.clinic
            if original_clinic != user:
                return Response({'error': 'Only the original clinic can subcontract this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            # Проверяем что договор еще не передан на субподряд
            if contract.is_subcontracted:
                return Response({'error': 'Contract is already subcontracted'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Получаем клинику-субподрядчика
            subcontractor_clinic = None
            if subcontractor_clinic_id:
                subcontractor_clinic = User.objects.get(id=subcontractor_clinic_id, role='clinic')
            elif subcontractor_clinic_bin:
                # Поиск по БИН/ИНН
                bin_normalized = ''.join(str(subcontractor_clinic_bin).strip().split())
                clinics = User.objects.filter(role='clinic')
                for clinic in clinics:
                    reg_data = clinic.registration_data or {}
                    clinic_bin = str(reg_data.get('bin', '')).strip()
                    clinic_inn = str(reg_data.get('inn', '')).strip()
                    clinic_bin_normalized = ''.join(clinic_bin.split())
                    clinic_inn_normalized = ''.join(clinic_inn.split())
                    if clinic_bin_normalized == bin_normalized or clinic_inn_normalized == bin_normalized:
                        subcontractor_clinic = clinic
                        break
                
                if not subcontractor_clinic:
                    # Клиника не найдена, но можем создать запрос для незарегистрированной
                    pass
            
            if not subcontractor_clinic and not subcontractor_clinic_phone:
                return Response({'error': 'subcontractor_clinic_id or subcontractor_clinic_bin with subcontractor_clinic_phone is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что не передаем самому себе
            if subcontractor_clinic and subcontractor_clinic == original_clinic:
                return Response({'error': 'Cannot subcontract to yourself'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Обновляем договор
            contract.is_subcontracted = True
            contract.subcontract_status = 'pending'
            contract.original_clinic = original_clinic
            contract.subcontract_amount = subcontract_amount
            if subcontractor_clinic:
                contract.subcontractor_clinic = subcontractor_clinic
            contract.subcontracted_at = timezone.now()
            # Меняем clinic на субподрядчика, чтобы он видел договор как свой
            if subcontractor_clinic:
                contract.clinic = subcontractor_clinic
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            subcontractor_name = subcontractor_clinic.registration_data.get('name', '') if subcontractor_clinic and subcontractor_clinic.registration_data else (subcontractor_clinic_bin or 'Клиника')
            ContractHistory.objects.create(
                contract=contract,
                action='subcontracted',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=contract.status,
                new_status=contract.status,
                comment=f'Договор передан на субподряд клинике: {subcontractor_name}. Сумма субподряда: {subcontract_amount} тенге'
            )
            
            # Отправляем уведомление
            if subcontractor_clinic_phone:
                self._send_subcontract_notification(contract, subcontractor_clinic_phone, subcontractor_clinic_bin or '', subcontract_amount, subcontractor_clinic is not None)
            
            # Обновляем сериализатор для возврата обновленных данных
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response({'error': 'User or subcontractor clinic not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc() if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def accept_subcontract(self, request, pk=None):
        """Принятие договора на субподряд клиникой-субподрядчиком"""
        try:
            contract = self.get_object()
            user_id = request.data.get('user_id')
            
            if not user_id:
                return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что пользователь является субподрядчиком
            user = User.objects.get(id=user_id)
            if user.role != 'clinic':
                return Response({'error': 'Only clinic can accept subcontract'}, status=status.HTTP_403_FORBIDDEN)
            
            if not contract.is_subcontracted or contract.subcontractor_clinic != user:
                return Response({'error': 'You are not the subcontractor for this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            if contract.subcontract_status != 'pending':
                return Response({'error': f'Contract subcontract status is {contract.subcontract_status}, not pending'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Обновляем статус
            contract.subcontract_status = 'accepted'
            contract.subcontract_accepted_at = timezone.now()
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='subcontracted',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=contract.status,
                new_status=contract.status,
                comment='Субподряд принят клиникой-субподрядчиком'
            )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc() if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def reject_subcontract(self, request, pk=None):
        """Отклонение договора на субподряд клиникой-субподрядчиком"""
        try:
            contract = self.get_object()
            user_id = request.data.get('user_id')
            reason = request.data.get('reason', '')
            
            if not user_id:
                return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что пользователь является субподрядчиком
            user = User.objects.get(id=user_id)
            if user.role != 'clinic':
                return Response({'error': 'Only clinic can reject subcontract'}, status=status.HTTP_403_FORBIDDEN)
            
            if not contract.is_subcontracted or contract.subcontractor_clinic != user:
                return Response({'error': 'You are not the subcontractor for this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            if contract.subcontract_status != 'pending':
                return Response({'error': f'Contract subcontract status is {contract.subcontract_status}, not pending'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Обновляем статус
            contract.subcontract_status = 'rejected'
            contract.subcontract_rejected_at = timezone.now()
            contract.subcontract_rejection_reason = reason
            # Возвращаем clinic обратно к оригинальной клинике
            if contract.original_clinic:
                contract.clinic = contract.original_clinic
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='subcontracted',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=contract.status,
                new_status=contract.status,
                comment=f'Субподряд отклонен клиникой-субподрядчиком. Причина: {reason}' if reason else 'Субподряд отклонен клиникой-субподрядчиком'
            )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc() if settings.DEBUG else None
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def _send_contract_rejection_notification(self, contract, phone, reason):
        """Отправка уведомления клинике об отклонении договора через WhatsApp"""
        try:
            from django.conf import settings
            import requests
            
            # Форматируем телефон
            formatted_phone = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            if not formatted_phone.startswith('7'):
                formatted_phone = '7' + formatted_phone
            
            # Получаем имя работодателя
            employer_name = 'Работодатель'
            if contract.employer:
                employer_name = contract.employer.registration_data.get('name', 'Работодатель') if contract.employer.registration_data else 'Работодатель'
            elif contract.employer_name:
                employer_name = contract.employer_name
            
            # Формируем сообщение
            message = f"""Добрый день!

Договор №{contract.contract_number} от {contract.contract_date} был отклонен работодателем.

Причина отклонения: {reason}

Вы можете отредактировать договор и отправить его повторно на согласование.

Для просмотра договора перейдите по ссылке:
{settings.FRONTEND_URL if hasattr(settings, 'FRONTEND_URL') else 'http://localhost:3000'}/dashboard/clinic/contracts"""
            
            # Отправляем через Green API
            url = f"{settings.GREEN_API_URL}/waInstance{settings.GREEN_API_ID_INSTANCE}/sendMessage/{settings.GREEN_API_TOKEN}"
            payload = {
                "chatId": f"{formatted_phone}@c.us",
                "message": message
            }
            response = requests.post(url, json=payload, timeout=10)
            
            if response.status_code != 200:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to send contract rejection notification: {response.status_code}")
        except Exception as e:
            # Логируем ошибку, но не прерываем отклонение договора
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send contract rejection notification: {str(e)}")
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Согласование договора"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            now = timezone.now()
            
            # Проверяем, является ли пользователь работодателем
            is_employer = False
            if user.role == 'employer':
                # Проверяем по employer или по БИН
                if contract.employer and user == contract.employer:
                    is_employer = True
                elif contract.employer_bin:
                    # Проверяем по БИН
                    reg_data = user.registration_data or {}
                    user_bin = reg_data.get('bin') or reg_data.get('inn')
                    if user_bin:
                        # Нормализуем БИН для сравнения
                        contract_bin_normalized = ''.join(str(contract.employer_bin).strip().split())
                        user_bin_normalized = ''.join(str(user_bin).strip().split())
                        if contract_bin_normalized == user_bin_normalized:
                            is_employer = True
                            # Если employer был None, устанавливаем его
                            if not contract.employer:
                                contract.employer = user
            
            if is_employer:
                contract.approved_by_employer_at = now
                # Сохраняем текущий статус для проверки
                current_status = contract.status
                
                # Если клиника уже подписала, статус становится 'active' (действует)
                if contract.approved_by_clinic_at:
                    contract.status = 'active'
                # Если статус 'pending_approval', и работодатель подписывает, статус становится 'approved'
                elif current_status == 'pending_approval':
                    contract.status = 'approved'
                # Если статус 'draft', меняем на 'pending_approval'
                elif current_status == 'draft':
                    contract.status = 'pending_approval'
                # В любом другом случае, если работодатель подписывает, статус становится 'approved'
                else:
                    contract.status = 'approved'
            elif user == contract.clinic:
                contract.approved_by_clinic_at = now
                # Сохраняем текущий статус для проверки
                current_status = contract.status
                
                # Если работодатель уже подписал, статус становится 'active' (действует)
                if contract.approved_by_employer_at:
                    contract.status = 'active'
                # Если статус 'pending_approval', и клиника подписывает, статус становится 'approved'
                elif current_status == 'pending_approval':
                    contract.status = 'approved'
                # Если статус 'draft', меняем на 'pending_approval'
                elif current_status == 'draft':
                    contract.status = 'pending_approval'
                # В любом другом случае, если клиника подписывает, статус становится 'approved'
                else:
                    contract.status = 'approved'
            else:
                return Response({'error': 'User is not authorized to approve this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            # Сохраняем изменения
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='approved',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=current_status,
                new_status=contract.status,
                comment=f'Договор согласован {"работодателем" if is_employer else "клиникой"}'
            )
            
            # Обновляем объект из базы данных, чтобы получить актуальный статус
            contract.refresh_from_db()
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Отклонение договора работодателем с указанием причины"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        reason = request.data.get('reason', '')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not reason:
            return Response({'error': 'reason is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Проверяем, является ли пользователь работодателем
            is_employer = False
            if user.role == 'employer':
                if contract.employer and user == contract.employer:
                    is_employer = True
                elif contract.employer_bin:
                    reg_data = user.registration_data or {}
                    user_bin = reg_data.get('bin') or reg_data.get('inn')
                    if user_bin:
                        contract_bin_normalized = ''.join(str(contract.employer_bin).strip().split())
                        user_bin_normalized = ''.join(str(user_bin).strip().split())
                        if contract_bin_normalized == user_bin_normalized:
                            is_employer = True
                            if not contract.employer:
                                contract.employer = user
            
            if not is_employer:
                return Response({'error': 'Only employer can reject contract'}, status=status.HTTP_403_FORBIDDEN)
            
            old_status = contract.status
            contract.status = 'rejected'
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='rejected',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=old_status,
                new_status='rejected',
                comment=reason
            )
            
            # Отправляем уведомление клинике о отклонении договора
            if contract.clinic:
                # Пытаемся получить телефон из разных источников
                clinic_phone = ''
                if contract.clinic.registration_data:
                    clinic_phone = contract.clinic.registration_data.get('phone', '')
                # Если нет в registration_data, проверяем другие поля
                if not clinic_phone and hasattr(contract.clinic, 'phone'):
                    clinic_phone = contract.clinic.phone
                
                if clinic_phone:
                    self._send_contract_rejection_notification(contract, clinic_phone, reason)
                else:
                    # Логируем, если не удалось найти телефон
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Could not find phone for clinic {contract.clinic.id} to send rejection notification")
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['patch'])
    def update_contract(self, request, pk=None):
        """Обновление договора клиникой (только в статусах draft или rejected)"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Только клиника может редактировать
            if user != contract.clinic:
                return Response({'error': 'Only clinic can update contract'}, status=status.HTTP_403_FORBIDDEN)
            
            # Можно редактировать только в статусах draft или rejected
            if contract.status not in ['draft', 'rejected']:
                return Response({'error': f'Cannot update contract in status {contract.status}'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Сохраняем старые значения для истории
            old_values = {
                'contract_number': contract.contract_number,
                'contract_date': str(contract.contract_date),
                'amount': str(contract.amount),
                'people_count': contract.people_count,
                'execution_date': str(contract.execution_date),
            }
            
            # Обновляем поля
            changes = {}
            if 'contract_number' in request.data and request.data['contract_number'] != contract.contract_number:
                changes['contract_number'] = {'old': contract.contract_number, 'new': request.data['contract_number']}
                contract.contract_number = request.data['contract_number']
            
            if 'contract_date' in request.data and request.data['contract_date'] != str(contract.contract_date):
                changes['contract_date'] = {'old': str(contract.contract_date), 'new': request.data['contract_date']}
                contract.contract_date = request.data['contract_date']
            
            if 'amount' in request.data and float(request.data['amount']) != float(contract.amount):
                changes['amount'] = {'old': str(contract.amount), 'new': str(request.data['amount'])}
                contract.amount = request.data['amount']
            
            if 'people_count' in request.data and int(request.data['people_count']) != contract.people_count:
                changes['people_count'] = {'old': contract.people_count, 'new': int(request.data['people_count'])}
                contract.people_count = request.data['people_count']
            
            if 'execution_date' in request.data and request.data['execution_date'] != str(contract.execution_date):
                changes['execution_date'] = {'old': str(contract.execution_date), 'new': request.data['execution_date']}
                contract.execution_date = request.data['execution_date']
            
            if 'notes' in request.data:
                if request.data['notes'] != contract.notes:
                    changes['notes'] = {'old': contract.notes, 'new': request.data['notes']}
                contract.notes = request.data['notes']
            
            contract.save()
            
            # Создаем запись в истории, если были изменения
            if changes:
                user_name = user.registration_data.get('name', '') if user.registration_data else ''
                ContractHistory.objects.create(
                    contract=contract,
                    action='updated',
                    user=user,
                    user_role=user.role,
                    user_name=user_name,
                    old_status=contract.status,
                    new_status=contract.status,
                    changes=changes,
                    comment='Договор обновлен клиникой'
                )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def resend_for_approval(self, request, pk=None):
        """Повторная отправка договора на согласование после доработки"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        comment = request.data.get('comment', '')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Только клиника может отправлять на согласование
            if user != contract.clinic:
                return Response({'error': 'Only clinic can resend contract'}, status=status.HTTP_403_FORBIDDEN)
            
            # Можно отправить только из статусов draft или rejected
            if contract.status not in ['draft', 'rejected']:
                return Response({'error': f'Cannot resend contract in status {contract.status}'}, status=status.HTTP_400_BAD_REQUEST)
            
            old_status = contract.status
            contract.status = 'pending_approval'
            contract.sent_at = timezone.now()
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='resent_for_approval' if old_status == 'rejected' else 'sent_for_approval',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=old_status,
                new_status='pending_approval',
                comment=comment or ('Договор повторно отправлен на согласование после доработки' if old_status == 'rejected' else 'Договор отправлен на согласование')
            )
            
            # Отправляем уведомление работодателю
            if contract.employer_phone:
                self._send_contract_notification(contract, contract.employer_phone, contract.employer_bin or '')
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Получение истории изменений договора"""
        contract = self.get_object()
        history_records = contract.history.all()
        serializer = ContractHistorySerializer(history_records, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Отправка договора"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            if user not in [contract.employer, contract.clinic]:
                return Response({'error': 'User is not authorized to send this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            if contract.status != 'approved':
                return Response({'error': 'Contract must be approved before sending'}, status=status.HTTP_400_BAD_REQUEST)
            
            contract.status = 'sent'
            contract.sent_at = timezone.now()
            contract.save()
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """Начало исполнения договора (переход в статус 'в процессе')"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            if user not in [contract.employer, contract.clinic]:
                return Response({'error': 'User is not authorized to execute this contract'}, status=status.HTTP_403_FORBIDDEN)
            
            if contract.status != 'active':
                return Response({'error': 'Contract must be active (действует) before starting execution'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Переводим договор в статус "в процессе исполнения"
            contract.status = 'in_progress'
            contract.executed_at = timezone.now()
            contract.save()
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['get'])
    def check_completion(self, request, pk=None):
        """Проверка статуса исполнения договора на основе завершенных медосмотров"""
        contract = self.get_object()
        
        # Получаем общее количество людей по договору
        total_people = contract.people_count
        
        if total_people == 0:
            return Response({
                'status': contract.status,
                'completed': 0,
                'total': 0,
                'percentage': 0,
                'message': 'Количество людей не указано'
            }, status=status.HTTP_200_OK)
        
        # Подсчитываем количество завершенных медосмотров
        # Медосмотр считается завершенным, если есть заключение профпатолога (Expertise с final_verdict)
        completed_count = 0
        
        # Получаем всех сотрудников работодателя по договору
        if contract.employer:
            # Получаем все экспертизы для сотрудников работодателя
            expertises = Expertise.objects.filter(
                user=contract.clinic,
                final_verdict__isnull=False
            )
            
            # Подсчитываем уникальных пациентов с завершенными экспертизами
            completed_patients = set()
            for expertise in expertises:
                # Проверяем, что пациент относится к работодателю по договору
                employee = ContingentEmployee.objects.filter(
                    user=contract.employer,
                    iin=expertise.iin
                ).first()
                if employee:
                    completed_patients.add(expertise.patient_id)
            
            completed_count = len(completed_patients)
        
        # Вычисляем процент выполнения
        percentage = (completed_count / total_people) * 100 if total_people > 0 else 0
        
        # Обновляем статус договора на основе прогресса
        old_status = contract.status
        if percentage == 100:
            contract.status = 'executed'
        elif percentage > 0 and contract.status == 'in_progress':
            contract.status = 'partially_executed'
        
        if old_status != contract.status:
            contract.save()
        
        return Response({
            'status': contract.status,
            'completed': completed_count,
            'total': total_people,
            'percentage': round(percentage, 2),
            'message': f'Завершено {completed_count} из {total_people} медосмотров'
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def upload_scan(self, request, pk=None):
        """Загрузка скан-файла договора"""
        contract = self.get_object()
        scan_file = request.FILES.get('file')
        
        if not scan_file:
            return Response({'error': 'File is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # В реальном приложении здесь должна быть логика сохранения файла
        # Пока сохраняем только имя файла в JSON
        scan_files = contract.scan_files if contract.scan_files else []
        scan_files.append({
            'filename': scan_file.name,
            'uploaded_at': timezone.now().isoformat(),
            'size': scan_file.size,
        })
        contract.scan_files = scan_files
        contract.save()
        
        serializer = self.get_serializer(contract)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def mark_executed_by_clinic(self, request, pk=None):
        """Клиника отмечает договор как исполненный (полностью или частично)"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        execution_type = request.data.get('execution_type')  # 'full' или 'partial'
        execution_notes = request.data.get('execution_notes', '')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not execution_type or execution_type not in ['full', 'partial']:
            return Response({'error': 'execution_type is required and must be "full" or "partial"'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Только клиника может отметить исполнение
            if user != contract.clinic:
                return Response({'error': 'Only clinic can mark contract as executed'}, status=status.HTTP_403_FORBIDDEN)
            
            # Договор должен быть в статусе 'active' или 'in_progress'
            if contract.status not in ['active', 'in_progress', 'partially_executed']:
                return Response({'error': f'Cannot mark contract as executed in status {contract.status}'}, status=status.HTTP_400_BAD_REQUEST)
            
            old_status = contract.status
            
            # Обновляем поля
            contract.execution_type = execution_type
            contract.executed_by_clinic_at = timezone.now()
            contract.execution_notes = execution_notes
            
            # Меняем статус на "ожидает подтверждения работодателем"
            # Используем существующий статус 'partially_executed' для частичного исполнения
            # и 'executed' для полного, но добавляем флаг что ждем подтверждения
            if execution_type == 'partial':
                contract.status = 'partially_executed'
            else:
                contract.status = 'executed'
            
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            execution_type_label = 'Полное исполнение' if execution_type == 'full' else 'Частичное исполнение'
            ContractHistory.objects.create(
                contract=contract,
                action='executed',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=old_status,
                new_status=contract.status,
                comment=f'Клиника отметила договор как исполненный. Тип: {execution_type_label}. {execution_notes}'
            )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def confirm_execution_by_employer(self, request, pk=None):
        """Работодатель подтверждает исполнение договора"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Проверяем, является ли пользователь работодателем
            is_employer = False
            if user.role == 'employer':
                if contract.employer and user == contract.employer:
                    is_employer = True
                elif contract.employer_bin:
                    reg_data = user.registration_data or {}
                    user_bin = reg_data.get('bin') or reg_data.get('inn')
                    if user_bin:
                        contract_bin_normalized = ''.join(str(contract.employer_bin).strip().split())
                        user_bin_normalized = ''.join(str(user_bin).strip().split())
                        if contract_bin_normalized == user_bin_normalized:
                            is_employer = True
                            if not contract.employer:
                                contract.employer = user
            
            if not is_employer:
                return Response({'error': 'Only employer can confirm execution'}, status=status.HTTP_403_FORBIDDEN)
            
            # Договор должен быть отмечен клиникой как исполненный
            if not contract.executed_by_clinic_at:
                return Response({'error': 'Contract must be marked as executed by clinic first'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что еще не подтверждено и не отклонено
            if contract.confirmed_by_employer_at:
                return Response({'error': 'Contract execution already confirmed by employer'}, status=status.HTTP_400_BAD_REQUEST)
            
            if contract.employer_rejection_reason:
                return Response({'error': 'Contract execution was rejected by employer'}, status=status.HTTP_400_BAD_REQUEST)
            
            old_status = contract.status
            
            # Подтверждаем исполнение
            contract.confirmed_by_employer_at = timezone.now()
            # Статус остается 'executed' или 'partially_executed' в зависимости от типа исполнения
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='executed',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=old_status,
                new_status=contract.status,
                comment='Работодатель подтвердил исполнение договора'
            )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def reject_execution_by_employer(self, request, pk=None):
        """Работодатель отклоняет исполнение договора с указанием причины"""
        contract = self.get_object()
        user_id = request.data.get('user_id')
        rejection_reason = request.data.get('rejection_reason', '')
        
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not rejection_reason:
            return Response({'error': 'rejection_reason is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
            
            # Проверяем, является ли пользователь работодателем
            is_employer = False
            if user.role == 'employer':
                if contract.employer and user == contract.employer:
                    is_employer = True
                elif contract.employer_bin:
                    reg_data = user.registration_data or {}
                    user_bin = reg_data.get('bin') or reg_data.get('inn')
                    if user_bin:
                        contract_bin_normalized = ''.join(str(contract.employer_bin).strip().split())
                        user_bin_normalized = ''.join(str(user_bin).strip().split())
                        if contract_bin_normalized == user_bin_normalized:
                            is_employer = True
                            if not contract.employer:
                                contract.employer = user
            
            if not is_employer:
                return Response({'error': 'Only employer can reject execution'}, status=status.HTTP_403_FORBIDDEN)
            
            # Договор должен быть отмечен клиникой как исполненный
            if not contract.executed_by_clinic_at:
                return Response({'error': 'Contract must be marked as executed by clinic first'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Проверяем что еще не подтверждено и не отклонено
            if contract.confirmed_by_employer_at:
                return Response({'error': 'Contract execution already confirmed by employer'}, status=status.HTTP_400_BAD_REQUEST)
            
            if contract.employer_rejection_reason:
                return Response({'error': 'Contract execution already rejected by employer'}, status=status.HTTP_400_BAD_REQUEST)
            
            old_status = contract.status
            
            # Отклоняем исполнение
            contract.employer_rejection_reason = rejection_reason
            # Возвращаем статус обратно в 'in_progress'
            contract.status = 'in_progress'
            contract.save()
            
            # Создаем запись в истории
            user_name = user.registration_data.get('name', '') if user.registration_data else ''
            ContractHistory.objects.create(
                contract=contract,
                action='rejected',
                user=user,
                user_role=user.role,
                user_name=user_name,
                old_status=old_status,
                new_status='in_progress',
                comment=f'Работодатель отклонил исполнение договора. Причина: {rejection_reason}'
            )
            
            serializer = self.get_serializer(contract)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)



# Health check endpoint для blue-green deployment
@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint для мониторинга состояния сервиса.
    Используется в blue-green deployment для проверки готовности сервиса.
    """
    try:
        # Проверяем подключение к БД
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        
        # Проверяем что можем получить данные из БД
        user_count = User.objects.count()
        
        return Response({
            'status': 'healthy',
            'timestamp': timezone.now().isoformat(),
            'database': 'connected',
            'users_count': user_count,
            'deployment_color': os.environ.get('DEPLOYMENT_COLOR', 'unknown')
        }, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({
            'status': 'unhealthy',
            'timestamp': timezone.now().isoformat(),
            'error': str(e),
            'deployment_color': os.environ.get('DEPLOYMENT_COLOR', 'unknown')
        }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
