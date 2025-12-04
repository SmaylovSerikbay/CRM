"""
Скрипт для проверки структуры Excel файла
"""
from openpyxl import load_workbook
import sys

def check_excel_file(filename):
    print(f"\n{'='*60}")
    print(f"Проверка файла: {filename}")
    print(f"{'='*60}\n")
    
    try:
        wb = load_workbook(filename)
        
        # Список всех листов
        print(f"📋 Листы в файле:")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            state = sheet.sheet_state
            print(f"  - {sheet_name} (состояние: {state})")
        
        # Проверка основного листа
        main_sheet = wb.active
        print(f"\n📊 Активный лист: {main_sheet.title}")
        
        # Ищем лист "Список контингента"
        contingent_sheet = None
        for sheet_name in wb.sheetnames:
            if "контингент" in sheet_name.lower():
                contingent_sheet = wb[sheet_name]
                break
        
        if contingent_sheet and contingent_sheet != main_sheet:
            print(f"   (Проверяем Data Validation на листе: {contingent_sheet.title})")
            check_sheet = contingent_sheet
        else:
            check_sheet = main_sheet
        
        # Проверка Data Validation
        print(f"\n🔍 Data Validation правила:")
        if hasattr(check_sheet, 'data_validations') and check_sheet.data_validations:
            if len(check_sheet.data_validations.dataValidation) > 0:
                for idx, dv in enumerate(check_sheet.data_validations.dataValidation, 1):
                    print(f"\n  Правило #{idx}:")
                    print(f"    Тип: {dv.type}")
                    print(f"    Формула: {dv.formula1}")
                    print(f"    Диапазоны: {dv.sqref}")
                    print(f"    Показывать выпадающий список: {dv.showDropDown}")
                    if dv.prompt:
                        print(f"    Подсказка: {dv.prompt}")
                    if dv.error:
                        print(f"    Сообщение об ошибке: {dv.error}")
            else:
                print("  ❌ Data Validation правила не найдены!")
        else:
            print("  ❌ Data Validation не настроена!")
        
        # Проверка справочных листов
        print(f"\n📚 Справочные листы:")
        for ref_sheet_name in ['Ref_Gender', 'Ref_Harm']:
            if ref_sheet_name in wb.sheetnames:
                ref_sheet = wb[ref_sheet_name]
                values = []
                for row in ref_sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                    if row[0]:
                        values.append(row[0])
                print(f"\n  {ref_sheet_name}:")
                print(f"    Состояние: {ref_sheet.sheet_state}")
                print(f"    Первые значения: {values[:5]}")
                if len(values) > 5:
                    print(f"    ... всего {len(values)} значений")
            else:
                print(f"\n  ❌ {ref_sheet_name} не найден!")
        
        wb.close()
        print(f"\n{'='*60}")
        print("✅ Проверка завершена")
        print(f"{'='*60}\n")
        
    except Exception as e:
        print(f"❌ Ошибка при проверке файла: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import sys
    import os
    
    # Если передан аргумент, проверяем конкретный файл
    if len(sys.argv) > 1:
        filename = sys.argv[1]
        if os.path.exists(filename):
            check_excel_file(filename)
        else:
            print(f"\n⚠️  Файл {filename} не найден\n")
    else:
        # Проверяем последние созданные файлы
        import glob
        files = glob.glob("test_contingent_template_*.xlsx") + glob.glob("template_with_instructions_*.xlsx")
        files.sort(key=os.path.getmtime, reverse=True)
        
        if files:
            print(f"\n🔍 Найдено {len(files)} файлов. Проверяем последний...\n")
            check_excel_file(files[0])
        else:
            print("\n⚠️  Файлы не найдены\n")
